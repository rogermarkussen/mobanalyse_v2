from pathlib import Path

import matplotlib.pyplot as plt
import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "mobil.parquet"
FIGURE_DIR = ROOT / "dist" / "assets" / "figures"
EXPORT_DIR = ROOT / "dist" / "assets" / "exports"

GROUP_COLORS = {
    "Telenor": "#00B0F0",
    "Telia": "#9900CC",
    "Lyse Tele (Ice)": "#ffc000",
    "Øvrige": "#00b050",
}
GROUP_ORDER = ["Telenor", "Telia", "Lyse Tele (Ice)", "Øvrige"]


def percent(value):
    return f"{value:.1f} %".replace(".", ",")


def provider_group():
    name = pl.col("fusnavn").str.to_lowercase()
    return (
        pl.when(name.str.contains("telenor"))
        .then(pl.lit("Telenor"))
        .when(name.str.contains("telia"))
        .then(pl.lit("Telia"))
        .when(
            name.is_in(["ice communication norge", "lyse tele"])
            | name.str.contains("lyse")
            | name.str.contains("ice")
        )
        .then(pl.lit("Lyse Tele (Ice)"))
        .otherwise(pl.lit("Øvrige"))
    )


def save_excel(data, filename, title):
    index_columns = [
        column for column in ["serie", "ms", "tilbyder"] if column in data.columns
    ]
    years = sorted(data["ar"].unique().to_list())
    export_data = (
        data.with_columns(
            (pl.col("markedsandel") / 100).alias("markedsandel"),
            pl.col("ar").cast(pl.Utf8),
        )
        .pivot(
            values="markedsandel",
            index=index_columns,
            on="ar",
            aggregate_function="first",
        )
        .select(index_columns + [str(year) for year in years])
        .sort(index_columns)
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=14, color="0B2B66")
    sheet.append([])
    sheet.append(export_data.columns)
    for row in export_data.iter_rows():
        sheet.append(list(row))

    for cell in sheet[3]:
        cell.fill = PatternFill("solid", fgColor="0B2B66")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
        for cell in row:
            if sheet.cell(row=3, column=cell.column).value in [str(year) for year in years]:
                cell.number_format = "0.0%"
    for column_index in range(1, sheet.max_column + 1):
        sheet.column_dimensions[
            sheet.cell(row=3, column=column_index).column_letter
        ].width = 14

    workbook.save(EXPORT_DIR / filename)


def plot_lines(data, filename, colors, order, upper_y, tick_step=10, legend_cols=None):
    fig, ax = plt.subplots(figsize=(7.2, 4.8), dpi=120)
    for provider in order:
        provider_data = data.filter(pl.col("tilbyder") == provider).sort("ar")
        if order != GROUP_ORDER:
            provider_data = provider_data.filter(pl.col("markedsandel") > 0)
        if provider_data.is_empty():
            continue
        xs = provider_data["ar"].to_list()
        ys = provider_data["markedsandel"].to_list()
        ax.plot(
            xs,
            ys,
            color=colors[provider],
            linewidth=3.0 if len(order) <= 4 else 2.6,
            solid_capstyle="round",
            label=provider,
        )
        if order == GROUP_ORDER and len(ys) >= 2:
            ax.annotate(
                percent(ys[-2]),
                xy=(xs[-2], ys[-2]),
                xytext=(8, 7),
                textcoords="offset points",
                color="#404040",
                fontsize=9,
            )
        if order == GROUP_ORDER or ys[-1] > 0.05:
            ax.annotate(
                percent(ys[-1]),
                xy=(xs[-1], ys[-1]),
                xytext=(8, 0),
                textcoords="offset points",
                color="#404040",
                fontsize=8 if len(order) > 4 else 9,
            )

    ax.set_ylim(0, upper_y)
    ticks = [value * tick_step for value in range(int(upper_y / tick_step) + 1)]
    if tick_step < 1:
        ticks = [value / 2 for value in range(0, int(upper_y * 2) + 1)]
    ax.set_yticks(ticks)
    ax.set_yticklabels([percent(value) for value in ticks])
    ax.set_xticks(sorted(data["ar"].unique().to_list()))
    ax.grid(axis="y", color="#d9d9d9", linewidth=0.8)
    ax.grid(axis="x", visible=False)
    ax.spines[["top", "right", "left"]].set_visible(False)
    ax.spines["bottom"].set_color("#d9d9d9")
    ax.tick_params(axis="both", colors="#555555", length=0)
    ax.legend(
        loc="lower center",
        bbox_to_anchor=(0.5, -0.24 if len(order) <= 4 else -0.30),
        ncol=legend_cols or len(order),
        frameon=False,
        fontsize=9,
    )
    fig.tight_layout()
    fig.savefig(FIGURE_DIR / filename, bbox_inches="tight")
    plt.close(fig)


def linear_projection(data, periods_ahead=3):
    rows = []
    for provider in GROUP_ORDER:
        provider_data = data.filter(pl.col("tilbyder") == provider).sort("ar")
        if provider_data.height < 2:
            continue
        xs = [float(value) for value in provider_data["ar"].to_list()]
        ys = [float(value) for value in provider_data["markedsandel"].to_list()]
        x_mean = sum(xs) / len(xs)
        y_mean = sum(ys) / len(ys)
        denominator = sum((x - x_mean) ** 2 for x in xs)
        slope = (
            sum((x - x_mean) * (y - y_mean) for x, y in zip(xs, ys)) / denominator
            if denominator
            else 0
        )
        last_year = int(max(xs))
        last_value = ys[-1]
        for year in range(last_year, last_year + periods_ahead + 1):
            rows.append(
                {
                    "ar": year,
                    "tilbyder": provider,
                    "markedsandel": last_value + slope * (year - last_year),
                }
            )
    return pl.DataFrame(rows)


def projection_export(actual, projection):
    return (
        pl.concat(
            [
                actual.with_columns(pl.lit("Historikk").alias("serie")),
                projection.with_columns(
                    pl.lit("Lineær trend").alias("serie"),
                    pl.col("ar").cast(pl.Int32),
                ),
            ]
        )
        .select("serie", "ar", "tilbyder", "markedsandel")
        .sort("serie", "ar", "tilbyder")
    )


def plot_projection(actual, projection, filename, upper_y):
    fig, ax = plt.subplots(figsize=(7.8, 4.8), dpi=120)
    last_actual_year = int(actual["ar"].max())
    for provider in GROUP_ORDER:
        actual_provider = actual.filter(pl.col("tilbyder") == provider).sort("ar")
        projection_provider = projection.filter(pl.col("tilbyder") == provider).sort("ar")
        if actual_provider.is_empty() or projection_provider.is_empty():
            continue
        actual_xs = actual_provider["ar"].to_list()
        actual_ys = actual_provider["markedsandel"].to_list()
        projection_xs = projection_provider["ar"].to_list()
        projection_ys = projection_provider["markedsandel"].to_list()
        ax.plot(
            actual_xs,
            actual_ys,
            color=GROUP_COLORS[provider],
            linewidth=2.5,
            solid_capstyle="round",
            label=provider,
        )
        ax.plot(
            projection_xs,
            projection_ys,
            color=GROUP_COLORS[provider],
            linewidth=1.8,
            linestyle=":",
            label="_nolegend_",
        )
        latest_actual = actual_provider.filter(pl.col("ar") == last_actual_year)
        ax.annotate(
            percent(latest_actual["markedsandel"][0]),
            xy=(last_actual_year, latest_actual["markedsandel"][0]),
            xytext=(8, 7),
            textcoords="offset points",
            color="#404040",
            fontsize=8,
        )
        ax.annotate(
            percent(projection_ys[-1]),
            xy=(projection_xs[-1], projection_ys[-1]),
            xytext=(8, 0),
            textcoords="offset points",
            color="#202020",
            fontsize=8,
            bbox={
                "boxstyle": "square,pad=0.35",
                "facecolor": "white",
                "edgecolor": GROUP_COLORS[provider],
                "linewidth": 0.8,
            },
        )

    ax.set_ylim(0, upper_y)
    ax.set_yticks(range(0, upper_y + 1, 10))
    ax.set_yticklabels([percent(value) for value in range(0, upper_y + 1, 10)])
    ax.set_xticks(
        sorted(set(actual["ar"].unique().to_list() + projection["ar"].unique().to_list()))
    )
    ax.grid(axis="y", color="#d9d9d9", linewidth=0.8)
    ax.grid(axis="x", visible=False)
    ax.spines[["top", "right", "left"]].set_visible(False)
    ax.spines["bottom"].set_color("#d9d9d9")
    ax.tick_params(axis="both", colors="#555555", length=0)
    ax.legend(
        loc="lower center",
        bbox_to_anchor=(0.5, -0.26),
        ncol=4,
        frameon=False,
        fontsize=8,
    )
    fig.tight_layout()
    fig.savefig(FIGURE_DIR / filename, bbox_inches="tight")
    plt.close(fig)


def grouped_market_share(df, metric, segment=None):
    filters = [
        pl.col("dk") == "Mobiltelefoni",
        pl.col("hg") == metric,
        pl.col("tp") == "Sum",
        pl.col("sk") == "Sluttbruker",
        pl.col("delar") == "Helår",
    ]
    if segment is not None:
        filters.append(pl.col("ms") == segment)
    if metric == "Abonnement":
        filters.extend(
            [
                pl.col("n1").is_in(["Fakturert", "Kontantkort"]),
                pl.col("n2") == "Ingen",
            ]
        )
    group_cols = ["ar", "tilbyder"] if segment is None else ["ar", "ms", "tilbyder"]
    denominator_cols = ["ar"] if segment is None else ["ar", "ms"]
    return (
        df.filter(*filters)
        .with_columns(tilbyder=provider_group())
        .group_by(group_cols)
        .agg(pl.col("svar").sum().alias("verdi"))
        .with_columns(
            markedsandel=pl.col("verdi") * 100 / pl.col("verdi").sum().over(denominator_cols)
        )
        .select(group_cols + ["markedsandel"])
        .sort(group_cols)
        .collect()
    )


def challenger_share(df, segment, mapping):
    name = pl.col("fusnavn").str.to_lowercase()
    challenger = None
    for source, label in mapping.items():
        branch = pl.when(name == source).then(pl.lit(label))
        challenger = branch if challenger is None else challenger.when(name == source).then(pl.lit(label))
    challenger = challenger.otherwise(None)

    data = (
        df.filter(
            (pl.col("dk") == "Mobiltelefoni")
            & (pl.col("hg").is_in(["Abonnement", "Inntekter"]))
            & (pl.col("ms") == segment)
            & (pl.col("tp") == "Sum")
            & (pl.col("sk") == "Sluttbruker")
            & (pl.col("delar") == "Helår")
            & (
                (pl.col("hg") == "Inntekter")
                | (
                    (pl.col("n1").is_in(["Fakturert", "Kontantkort"]))
                    & (pl.col("n2") == "Ingen")
                )
            )
        )
        .with_columns(tilbyder=challenger, total=pl.col("svar").sum().over(["ar", "hg"]))
        .filter(pl.col("tilbyder").is_not_null())
        .group_by("ar", "hg", "tilbyder")
        .agg(pl.col("svar").sum().alias("verdi"), pl.col("total").first().alias("total"))
        .with_columns(markedsandel=pl.col("verdi") * 100 / pl.col("total"))
        .select("ar", "hg", "tilbyder", "markedsandel")
        .sort("hg", "ar", "tilbyder")
        .collect()
    )
    years = sorted(data["ar"].unique().to_list())
    metrics = ["Abonnement", "Inntekter"]
    providers = list(mapping.values())
    complete_grid = pl.DataFrame(
        [
            {"ar": year, "hg": metric, "tilbyder": provider}
            for year in years
            for metric in metrics
            for provider in providers
        ]
    )
    return (
        complete_grid.join(data, on=["ar", "hg", "tilbyder"], how="left")
        .with_columns(pl.col("markedsandel").fill_null(0))
        .sort("hg", "ar", "tilbyder")
    )


def add_nortel_from_powerpoint(data):
    nortel = pl.DataFrame(
        [
            {"ar": 2020, "hg": "Abonnement", "tilbyder": "Nortel", "markedsandel": 0.8098642261413776},
            {"ar": 2021, "hg": "Abonnement", "tilbyder": "Nortel", "markedsandel": 1.6072034025070056},
            {"ar": 2022, "hg": "Abonnement", "tilbyder": "Nortel", "markedsandel": 3.6328280585728845},
            {"ar": 2023, "hg": "Abonnement", "tilbyder": "Nortel", "markedsandel": 0.0},
            {"ar": 2024, "hg": "Abonnement", "tilbyder": "Nortel", "markedsandel": 0.0},
            {"ar": 2025, "hg": "Abonnement", "tilbyder": "Nortel", "markedsandel": 0.0},
            {"ar": 2020, "hg": "Inntekter", "tilbyder": "Nortel", "markedsandel": 0.6194134655626778},
            {"ar": 2021, "hg": "Inntekter", "tilbyder": "Nortel", "markedsandel": 1.9565108755375232},
            {"ar": 2022, "hg": "Inntekter", "tilbyder": "Nortel", "markedsandel": 3.799490531159897},
            {"ar": 2023, "hg": "Inntekter", "tilbyder": "Nortel", "markedsandel": 0.0},
            {"ar": 2024, "hg": "Inntekter", "tilbyder": "Nortel", "markedsandel": 0.0},
            {"ar": 2025, "hg": "Inntekter", "tilbyder": "Nortel", "markedsandel": 0.0},
        ],
        schema={
            "ar": pl.Int64,
            "hg": pl.Utf8,
            "tilbyder": pl.Utf8,
            "markedsandel": pl.Float64,
        },
    )
    return pl.concat([data.filter(pl.col("tilbyder") != "Nortel"), nortel]).sort(
        "hg", "ar", "tilbyder"
    )


def main():
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    df = pl.scan_parquet(DATA)

    ab_total = grouped_market_share(df, "Abonnement")
    rev_total = grouped_market_share(df, "Inntekter")
    plot_lines(ab_total, "figur-1-abonnement.png", GROUP_COLORS, GROUP_ORDER, 60)
    plot_lines(rev_total, "figur-1-omsetning.png", GROUP_COLORS, GROUP_ORDER, 60)
    save_excel(ab_total, "figur-1-abonnement.xlsx", "Figur 1 - Abonnement")
    save_excel(rev_total, "figur-1-omsetning.xlsx", "Figur 1 - Omsetning")

    ab_projection = linear_projection(ab_total)
    rev_projection = linear_projection(rev_total)
    plot_projection(ab_total, ab_projection, "figur-2-abonnement-trend.png", 60)
    plot_projection(rev_total, rev_projection, "figur-2-omsetning-trend.png", 60)
    save_excel(
        projection_export(ab_total, ab_projection),
        "figur-2-abonnement-trend.xlsx",
        "Figur 2 - Abonnement trend",
    )
    save_excel(
        projection_export(rev_total, rev_projection),
        "figur-2-omsetning-trend.xlsx",
        "Figur 2 - Omsetning trend",
    )

    ab_segment = grouped_market_share(df, "Abonnement", segment="Privat")
    ab_bedrift = grouped_market_share(df, "Abonnement", segment="Bedrift")
    plot_lines(ab_segment, "figur-3-privat.png", GROUP_COLORS, GROUP_ORDER, 50)
    plot_lines(ab_bedrift, "figur-3-bedrift.png", GROUP_COLORS, GROUP_ORDER, 60)
    save_excel(ab_segment, "figur-3-privat.xlsx", "Figur 3 - Privat")
    save_excel(ab_bedrift, "figur-3-bedrift.xlsx", "Figur 3 - Bedrift")

    rev_privat = grouped_market_share(df, "Inntekter", segment="Privat")
    rev_bedrift = grouped_market_share(df, "Inntekter", segment="Bedrift")
    plot_lines(rev_privat, "figur-4-privat-omsetning.png", GROUP_COLORS, GROUP_ORDER, 60)
    plot_lines(rev_bedrift, "figur-4-bedrift-omsetning.png", GROUP_COLORS, GROUP_ORDER, 70)
    save_excel(rev_privat, "figur-4-privat-omsetning.xlsx", "Figur 4 - Privat omsetning")
    save_excel(rev_bedrift, "figur-4-bedrift-omsetning.xlsx", "Figur 4 - Bedrift omsetning")

    private_mapping = {
        "fjordkraft mobil": "Fjordkraft",
        "chili mobil": "Chili mobil",
        "lycamobile norway ltd": "Lycamobile",
        "xplora mobile": "Xplora",
        "happybytes": "Happybytes",
        "plussmobil": "Plussmobil",
    }
    private_colors = {
        "Fjordkraft": "#385624",
        "Chili mobil": "#FF0000",
        "Lycamobile": "#002060",
        "Xplora": "#66FF99",
        "Happybytes": "#698ED0",
        "Plussmobil": "#A5A5A5",
    }
    private_order = list(private_colors)
    private = challenger_share(df, "Privat", private_mapping)
    for metric, suffix, upper in [
        ("Abonnement", "abonnement", 4),
        ("Inntekter", "omsetning", 3),
    ]:
        data = private.filter(pl.col("hg") == metric)
        plot_lines(
            data,
            f"figur-5-privat-utfordrere-{suffix}.png",
            private_colors,
            private_order,
            upper,
            tick_step=0.5,
            legend_cols=3,
        )
        save_excel(data, f"figur-5-privat-utfordrere-{suffix}.xlsx", f"Figur 5 - {suffix}")

    business_mapping = {
        "unifon": "Unifon",
        "nortel": "Nortel",
        "saga mobil": "Saga mobil",
        "smb mobil": "SMB mobil",
    }
    business_colors = {
        "Unifon": "#7F7F7F",
        "Nortel": "#64F600",
        "Saga mobil": "#C00000",
        "SMB mobil": "#548235",
    }
    business_order = list(business_colors)
    business = add_nortel_from_powerpoint(
        challenger_share(df, "Bedrift", business_mapping)
    )
    for metric, suffix, upper in [
        ("Abonnement", "abonnement", 8),
        ("Inntekter", "omsetning", 10),
    ]:
        data = business.filter(pl.col("hg") == metric)
        plot_lines(
            data,
            f"figur-6-bedrift-utfordrere-{suffix}.png",
            business_colors,
            business_order,
            upper,
            tick_step=1,
            legend_cols=4,
        )
        save_excel(data, f"figur-6-bedrift-utfordrere-{suffix}.xlsx", f"Figur 6 - {suffix}")


if __name__ == "__main__":
    main()
