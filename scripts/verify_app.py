from __future__ import annotations

import json
import math
import xml.etree.ElementTree as ET
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DIST = ROOT / "dist"
APP_DATA = DIST / "assets" / "data" / "app-data.json"
PPTX = ROOT / "sommervikar-endelig.pptx"

NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
}


def load_app_data() -> dict:
    if not APP_DATA.exists():
        raise AssertionError(f"Mangler {APP_DATA}")
    return json.loads(APP_DATA.read_text(encoding="utf-8"))


def chart_latest(chart_name: str, multiply: float = 1.0) -> dict[str, float]:
    with ZipFile(PPTX) as archive:
        root = ET.fromstring(archive.read(f"ppt/charts/{chart_name}.xml"))
    values: dict[str, float] = {}
    for series in root.findall(".//c:ser", NS):
        name_node = series.find(".//c:tx//c:v", NS)
        if name_node is None or not name_node.text:
            continue
        values[normalize_name(name_node.text)] = float(
            series.findall(".//c:val//c:numCache//c:pt/c:v", NS)[-1].text
        ) * multiply
    return values


def chart_points(chart_name: str, multiply: float = 1.0) -> dict[str, dict[int, float]]:
    with ZipFile(PPTX) as archive:
        root = ET.fromstring(archive.read(f"ppt/charts/{chart_name}.xml"))
    values: dict[str, dict[int, float]] = {}
    for series in root.findall(".//c:ser", NS):
        name_node = series.find(".//c:tx//c:v", NS)
        if name_node is None or not name_node.text:
            continue
        name = normalize_name(name_node.text)
        categories = [node.text or "" for node in series.findall(".//c:cat//c:strCache//c:pt/c:v", NS)]
        if not categories:
            categories = [
                node.text or "" for node in series.findall(".//c:cat//c:numCache//c:pt/c:v", NS)
            ]
        series_values = [
            float(node.text) * multiply
            for node in series.findall(".//c:val//c:numCache//c:pt/c:v", NS)
        ]
        for category, value in zip(categories, series_values):
            if category.isdigit():
                values.setdefault(name, {})[int(category)] = value
    return values


def slide_tables(slide_number: int) -> list[list[list[str]]]:
    with ZipFile(PPTX) as archive:
        root = ET.fromstring(archive.read(f"ppt/slides/slide{slide_number}.xml"))
    tables = []
    for table in root.findall(".//a:tbl", NS):
        rows = []
        for xml_row in table.findall("./a:tr", NS):
            cells = []
            for cell in xml_row.findall("./a:tc", NS):
                cells.append(" ".join(text.text or "" for text in cell.findall(".//a:t", NS)))
            rows.append(cells)
        tables.append(rows)
    return tables


def parse_percent(value: str) -> float:
    cleaned = value.replace("%", "").replace("\xa0", " ").strip().replace(",", ".")
    return float(cleaned.split()[0])


def parse_decimal(value: str) -> float:
    cleaned = value.replace("\xa0", " ").strip().replace(",", ".")
    return float(cleaned.split()[0])


def normalize_name(name: str) -> str:
    return {
        "Lyse (Ice)": "Lyse Tele (Ice)",
        "Happybytes ": "Happybytes",
        "ARPU bedriftsmarkedet": "Bedrift",
        "ARPU privatmarkedet": "Privat",
        "chili mobil": "Chili mobil",
        "happybytes": "Happybytes",
        "fjordkraft": "Fjordkraft",
        "plussmobil": "Plussmobil",
        "unifon": "Unifon",
    }.get(name, name)


def find_value(rows: list[dict], **match: object) -> float:
    for row in rows:
        if all(row.get(key) == value for key, value in match.items()):
            return float(row["value"])
    raise AssertionError(f"Mangler verdi for {match}")


def assert_close(label: str, current: float, baseline: float, tolerance: float) -> dict:
    delta = current - baseline
    if abs(delta) > tolerance:
        raise AssertionError(
            f"{label}: {current:.3f} er for langt fra PowerPoint {baseline:.3f} "
            f"(avvik {delta:.3f}, toleranse {tolerance:.3f})"
        )
    return {
        "label": label,
        "current": round(current, 4),
        "powerpoint": round(baseline, 4),
        "delta": round(delta, 4),
        "tolerance": tolerance,
    }


def assert_relative(label: str, current: float, baseline: float, rel: float, minimum: float) -> dict:
    tolerance = max(abs(baseline) * rel, minimum)
    return assert_close(label, current, baseline, tolerance)


def verify_static_files() -> None:
    required = [
        DIST / "index.html",
        DIST / "assets" / "app.js",
        DIST / "assets" / "styles.css",
        DIST / "assets" / "data" / "app-data.json",
        DIST / "data" / "mobil.parquet",
    ]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise AssertionError(f"Mangler statiske filer: {missing}")
    html = (DIST / "index.html").read_text(encoding="utf-8").lower()
    if "marimo" in html:
        raise AssertionError("dist/index.html inneholder fortsatt Marimo-skall")
    app_js = (DIST / "assets" / "app.js").read_text(encoding="utf-8")
    forbidden_ui = ["wholesalePeriod", "totalPeriod", "Uavhengig", '"Halvår"', "'Halvår'"]
    found = [term for term in forbidden_ui if term in app_js]
    if found:
        raise AssertionError(f"UI-koden inneholder gamle periode/uavhengig-valg: {found}")
    forbidden_public_text = [
        "mobil.parquet",
        "tilgangskjøper-valg.xlsx",
        "PowerPoint",
        "PPT-data",
        "Parquet",
        "dk =",
        "hg =",
        "tp =",
        "sk =",
        "fusnavn",
        "levnavn",
        "grnavn",
        "bygget fra",
    ]
    found_public = [term for term in forbidden_public_text if term in app_js]
    if found_public:
        raise AssertionError(f"UI-tekst inneholder interne referanser: {found_public}")


def verify_exports(data: dict) -> None:
    for export_id, info in data["exports"].items():
        for field in ["csvPath", "xlsxPath"]:
            path = DIST / info[field].replace("assets/", "assets/")
            if not path.exists():
                raise AssertionError(f"Mangler eksport {export_id}: {path}")
            if path.stat().st_size < 20:
                raise AssertionError(f"Eksport {path} er uventet liten")
    if "accessBuyers" in data or "access-buyers" in data["exports"]:
        raise AssertionError("Tilgangskjøper-tabellen er fortsatt pakket inn i grossistdata")


def verify_xlsx_layouts(data: dict) -> None:
    latest = int(data["metadata"]["latest_year"])
    for export_id, info in data["exports"].items():
        path = DIST / info["xlsxPath"]
        workbook = load_workbook(path, data_only=False, read_only=False)
        try:
            for sheet in workbook.worksheets:
                if sheet.max_row < 5 or sheet.max_column < 3:
                    raise AssertionError(f"{path.name}/{sheet.title} er for liten")
                if sheet["A1"].value in {None, ""}:
                    raise AssertionError(f"{path.name}/{sheet.title} mangler tittel i A1")
                if sheet.freeze_panes != "A5":
                    raise AssertionError(f"{path.name}/{sheet.title} mangler freeze pane A5")
                if not sheet.auto_filter.ref or not sheet.auto_filter.ref.startswith("A4:"):
                    raise AssertionError(f"{path.name}/{sheet.title} mangler filter fra rad 4")
                row_label = str(sheet["A4"].value or "").strip().lower()
                if row_label in {"ar", "år"}:
                    raise AssertionError(f"{path.name}/{sheet.title} bruker fortsatt år som radfelt")
                years = [
                    int(cell.value)
                    for cell in sheet[4][1:]
                    if str(cell.value).isdigit()
                ]
                if not years or max(years) < latest:
                    raise AssertionError(f"{path.name}/{sheet.title} mangler årskolonner")
            if export_id == "projection":
                sheet = workbook.worksheets[0]
                labels = {cell.value for cell in sheet[3] if cell.value}
                if not {"Historikk", "Lineær trend"}.issubset(labels):
                    raise AssertionError("Framskrivingseksporten mangler Historikk/Lineær trend-seksjoner")
                years = [int(cell.value) for cell in sheet[4][1:] if str(cell.value).isdigit()]
                if max(years) != latest + 3:
                    raise AssertionError("Framskrivingseksporten mangler tre trendår")
            if export_id == "market-share":
                sheet = workbook["Abonnement"]
                value = sheet["B5"].value
                if not isinstance(value, (int, float)) or not 0 <= value <= 1:
                    raise AssertionError("Markedsandeler eksporteres ikke som Excel-prosentverdier")
                if sheet["B5"].number_format != "0.0%":
                    raise AssertionError("Markedsandeler mangler prosentformat")
        finally:
            workbook.close()


def verify_sums(data: dict) -> None:
    for dataset_name in ["marketShare", "segmentShare"]:
        buckets: dict[tuple, float] = {}
        for row in data[dataset_name]:
            if dataset_name == "marketShare":
                key = (row["metric"], row["ar"])
            else:
                key = (row["metric"], row["segment"], row["ar"])
            buckets[key] = buckets.get(key, 0.0) + float(row["value"])
        bad = [(key, value) for key, value in buckets.items() if abs(value - 100) > 0.05]
        if bad:
            raise AssertionError(f"Andeler summerer ikke til 100 i {dataset_name}: {bad[:5]}")


def verify_projection(data: dict) -> None:
    latest = data["metadata"]["latest_year"]
    expected_last = latest + 3
    grouped: dict[tuple[str, str], set[int]] = {}
    for row in data["projection"]:
        grouped.setdefault((row["metric"], row["tilbyder"]), set()).add(int(row["ar"]))
    for key, years in grouped.items():
        if latest not in years or expected_last not in years:
            raise AssertionError(f"Projeksjon mangler forventede år for {key}: {sorted(years)}")


def verify_hel_year_and_wholesale_contract(data: dict) -> None:
    owners = data["order"].get("wholesaleOwners", [])
    if owners != ["Telenor", "Telia", "Lyse Tele (Ice)"]:
        raise AssertionError(f"Uventet grossistliste: {owners}")

    period_datasets = [
        "totals",
        "providerShareTrend",
        "providerSubscriptions",
        "wholesaleDefault",
        "grossistConcentrationDefault",
    ]
    for dataset_name in period_datasets:
        periods = {
            row.get("period")
            for row in data.get(dataset_name, [])
            if isinstance(row, dict) and "period" in row
        }
        if periods != {"Helår"}:
            raise AssertionError(f"{dataset_name} skal bare inneholde Helår, fant {periods}")

    template = data.get("wholesaleAssignmentTemplate", {})
    expected_years = {str(year) for year in range(2020, data["metadata"]["latest_year"] + 1)}
    if set(template) != expected_years:
        raise AssertionError(f"Grossistmal mangler år: {expected_years - set(template)}")

    valid_owners = set(owners)
    for row in data["providerSubscriptions"]:
        year = str(row["ar"])
        provider = row["provider"]
        owner = template.get(year, {}).get(provider)
        if owner not in valid_owners:
            raise AssertionError(f"Mangler gyldig grossist for {provider} i {year}: {owner}")

    expected_assignments = {
        ("2020", "lyse tele"): "Lyse Tele (Ice)",
        ("2022", "fjordkraft mobil"): "Telenor",
        ("2023", "fjordkraft mobil"): "Telia",
        ("2025", "lycamobile norway ltd"): "Telia",
    }
    for (year, provider), owner in expected_assignments.items():
        current = template.get(year, {}).get(provider)
        if current != owner:
            raise AssertionError(
                f"Excel-basert mapping er feil for {provider} i {year}: {current} != {owner}"
            )

    for row in data["wholesaleDefault"]:
        if row["grossist"] not in valid_owners:
            raise AssertionError(f"Ugyldig grossist i wholesaleDefault: {row}")


def verify_arpu_average_denominator(data: dict) -> None:
    for dataset_name in ["arpuSegment", "arpuProvider"]:
        rows = data.get(dataset_name, [])
        for row in rows:
            previous = row.get("abonnement_forrige_ar")
            current = float(row["abonnement_slutt_ar"])
            if previous is None:
                expected_abonnement = current
            else:
                expected_abonnement = (float(previous) + current) / 2.0
            if abs(float(row["abonnement"]) - expected_abonnement) > 0.01:
                raise AssertionError(
                    f"{dataset_name} bruker ikke snittabonnement for {row}: "
                    f"{row['abonnement']} != {expected_abonnement}"
                )
            expected_arpu = float(row["omsetning"]) * 1000 / float(row["abonnement"]) / 12
            if abs(float(row["value"]) - expected_arpu) > 0.0001:
                raise AssertionError(
                    f"{dataset_name} har feil ARPU-formel for {row}: "
                    f"{row['value']} != {expected_arpu}"
                )


def verify_wholesale_tables(data: dict) -> list[dict]:
    checks = []
    slide11 = slide_tables(11)[0]
    header = slide11[0]
    whole_year_columns = [
        (index, int(label))
        for index, label in enumerate(header)
        if label.isdigit() and 2020 <= int(label) <= 2024
    ]
    for row in slide11[1:]:
        if not row or not row[0]:
            continue
        provider = normalize_name(row[0])
        for column_index, year in whole_year_columns:
            current = find_value(
                data["wholesaleDefault"],
                period="Helår",
                ar=year,
                grossist=provider,
            )
            checks.append(
                assert_close(
                    f"Wholesale table {year}: {provider}",
                    current,
                    parse_percent(row[column_index]),
                    1.0,
                )
            )

    retail_table = slide_tables(12)[0]
    for whole_year_row in retail_table[2:]:
        if not whole_year_row or not whole_year_row[0].isdigit():
            continue
        year = int(whole_year_row[0])
        if not 2022 <= year <= 2024:
            continue
        current = next(
            row
            for row in data["concentration"]
            if row["metric"] == "Omsetning" and row["ar"] == year
        )
        checks.append(
            assert_close(
                f"Sluttbruker omsetning CR2 table {year}",
                float(current["cr2"]),
                parse_percent(whole_year_row[1]),
                0.5,
            )
        )
        checks.append(
            assert_close(
                f"Sluttbruker omsetning HHI table {year}",
                float(current["hhi"]),
                parse_decimal(whole_year_row[2]),
                0.01,
            )
        )

    grossist_table = slide_tables(12)[1]
    for whole_year_row in grossist_table[2:]:
        if not whole_year_row or not whole_year_row[0].isdigit():
            continue
        year = int(whole_year_row[0])
        if not 2022 <= year <= 2024:
            continue
        current = next(
            row
            for row in data["grossistConcentrationDefault"]
            if row["period"] == "Helår" and row["ar"] == year
        )
        checks.append(
            assert_close(
                f"Grossist CR2 table {year}",
                float(current["cr2"]),
                parse_percent(whole_year_row[1]),
                0.5,
            )
        )
        checks.append(
            assert_close(
                f"Grossist HHI table {year}",
                float(current["hhi"]),
                parse_decimal(whole_year_row[2]),
                0.01,
            )
        )
    return checks


def verify_powerpoint_baseline(data: dict) -> list[dict]:
    latest = data["metadata"]["latest_year"]
    checks: list[dict] = []

    chart_map = [
        (
            "chart1",
            "marketShare",
            {"metric": "Abonnement"},
            "tilbyder",
            100.0,
            1.0,
            3.5,
            "Abonnement total",
        ),
        (
            "chart2",
            "marketShare",
            {"metric": "Omsetning"},
            "tilbyder",
            100.0,
            1.0,
            3.5,
            "Omsetning total",
        ),
        (
            "chart5",
            "segmentShare",
            {"metric": "Abonnement", "segment": "Privat"},
            "tilbyder",
            100.0,
            1.0,
            4.5,
            "Abonnement privat",
        ),
        (
            "chart6",
            "segmentShare",
            {"metric": "Abonnement", "segment": "Bedrift"},
            "tilbyder",
            100.0,
            1.0,
            4.5,
            "Abonnement bedrift",
        ),
        (
            "chart7",
            "segmentShare",
            {"metric": "Omsetning", "segment": "Privat"},
            "tilbyder",
            100.0,
            1.0,
            5.0,
            "Omsetning privat",
        ),
        (
            "chart8",
            "segmentShare",
            {"metric": "Omsetning", "segment": "Bedrift"},
            "tilbyder",
            100.0,
            1.0,
            5.0,
            "Omsetning bedrift",
        ),
    ]

    for (
        chart,
        dataset,
        base_match,
        provider_key,
        multiplier,
        historical_tolerance,
        latest_tolerance,
        label,
    ) in chart_map:
        for provider, year_values in chart_points(chart, multiplier).items():
            for year, baseline in year_values.items():
                if 2020 <= year <= 2024:
                    current = find_value(
                        data[dataset],
                        **base_match,
                        **{provider_key: provider, "ar": year},
                    )
                    checks.append(
                        assert_close(
                            f"{label} {year}: {provider}",
                            current,
                            baseline,
                            historical_tolerance,
                        )
                    )
        ppt_values = chart_latest(chart, multiplier)
        for provider, baseline in ppt_values.items():
            current = find_value(
                data[dataset],
                **base_match,
                **{provider_key: provider, "ar": latest},
            )
            checks.append(assert_close(f"{label} siste dekkpunkt: {provider}", current, baseline, latest_tolerance))

    for segment, year_values in chart_points("chart13").items():
        for year, baseline in year_values.items():
            if 2020 <= year <= 2024:
                current = find_value(data["arpuSegment"], segment=segment, ar=year)
                checks.append(
                    assert_relative(f"ARPU {segment} {year}", current, baseline, rel=0.12, minimum=25)
                )

    for segment, baseline in chart_latest("chart13").items():
        current = find_value(data["arpuSegment"], segment=segment, ar=latest)
        checks.append(assert_relative(f"ARPU {segment}", current, baseline, rel=0.18, minimum=35))

    for provider, year_values in chart_points("chart16").items():
        for year, baseline in year_values.items():
            if 2020 <= year <= 2024:
                current = find_value(data["nokPerGbTotal"], tilbyder=provider, ar=year)
                checks.append(
                    assert_relative(
                        f"NOK per GB {provider} {year}",
                        current,
                        baseline,
                        rel=0.15,
                        minimum=5,
                    )
                )

    for provider, baseline in chart_latest("chart16").items():
        current = find_value(data["nokPerGbTotal"], tilbyder=provider, ar=latest)
        checks.append(
            assert_relative(f"NOK per GB {provider}", current, baseline, rel=0.35, minimum=8)
        )

    return checks


def summarize_baseline_checks(checks: list[dict]) -> dict:
    groups: dict[str, int] = {}
    max_tolerance_ratio = 0.0
    largest_relative_check = None
    for check in checks:
        group = str(check["label"]).split(":")[0].rsplit(" ", 1)[0]
        groups[group] = groups.get(group, 0) + 1
        tolerance = float(check["tolerance"])
        ratio = abs(float(check["delta"])) / tolerance if tolerance else 0.0
        if ratio > max_tolerance_ratio:
            max_tolerance_ratio = ratio
            largest_relative_check = check
    return {
        "check_count": len(checks),
        "max_tolerance_ratio": round(max_tolerance_ratio, 4),
        "largest_relative_check": largest_relative_check,
        "groups": dict(sorted(groups.items())),
    }


def verify_no_invalid_numbers(data: dict) -> None:
    for dataset_name, rows in data.items():
        if not isinstance(rows, list):
            continue
        for index, row in enumerate(rows):
            if not isinstance(row, dict):
                continue
            for key, value in row.items():
                if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
                    raise AssertionError(f"Ugyldig tall i {dataset_name}[{index}].{key}")


def main() -> None:
    verify_static_files()
    data = load_app_data()
    if data["metadata"]["latest_year"] < 2025:
        raise AssertionError("Datagrunnlaget mangler forventet 2025-årgang")
    verify_no_invalid_numbers(data)
    verify_exports(data)
    verify_xlsx_layouts(data)
    verify_sums(data)
    verify_projection(data)
    verify_hel_year_and_wholesale_contract(data)
    verify_arpu_average_denominator(data)
    baseline_checks = verify_powerpoint_baseline(data) + verify_wholesale_tables(data)

    report = {
        "status": "ok",
        "latest_year": data["metadata"]["latest_year"],
        "powerpoint_summary": summarize_baseline_checks(baseline_checks),
        "powerpoint_checks": baseline_checks,
    }
    report_path = DIST / "assets" / "data" / "verification.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        f"OK: statiske filer, andelssummer, projeksjoner og {len(baseline_checks)} "
        f"PowerPoint-sammenligninger er verifisert."
    )


if __name__ == "__main__":
    main()
