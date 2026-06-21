from __future__ import annotations

import csv
import json
import shutil
import subprocess
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DIST = ROOT / "dist"
DATA_DIR = DIST / "assets" / "data"
EXPORT_DIR = DIST / "assets" / "exports"
SOURCE_PARQUET = ROOT / "data" / "mobil.parquet"
ACCESS_BUYER_WORKBOOK = ROOT / "tilgangskjøper-valg.xlsx"
WHOLESALE_OWNERS = ["Telenor", "Telia", "Lyse Tele (Ice)"]


def run_duckdb(sql: str) -> list[dict]:
    if shutil.which("duckdb") is None:
        import duckdb

        connection = duckdb.connect()
        try:
            result = connection.execute(sql).fetchall()
            columns = [column[0] for column in connection.description or []]
            return [dict(zip(columns, row)) for row in result]
        finally:
            connection.close()

    result = subprocess.run(
        ["duckdb", "-json", "-c", sql],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
    )
    return json.loads(result.stdout or "[]")


def provider_group(column: str = "fusnavn", other: str = "'Øvrige'") -> str:
    name = f"lower({column})"
    return f"""
        CASE
            WHEN contains({name}, 'telenor') THEN 'Telenor'
            WHEN contains({name}, 'telia') THEN 'Telia'
            WHEN {name} IN ('ice communication norge', 'lyse tele')
                OR contains({name}, 'lyse')
                OR contains({name}, 'ice') THEN 'Lyse Tele (Ice)'
            ELSE {other}
        END
    """


def provider_price_group(column: str = "fusnavn", other: str = "NULL") -> str:
    name = f"lower({column})"
    return f"""
        CASE
            WHEN contains({name}, 'telenor') THEN 'Telenor'
            WHEN contains({name}, 'telia') THEN 'Telia'
            WHEN contains({name}, 'lyse') OR contains({name}, 'ice') THEN 'Ice'
            WHEN {name} = 'fjordkraft mobil' THEN 'Fjordkraft'
            WHEN {name} = 'chili mobil' THEN 'Chili mobil'
            WHEN {name} = 'plussmobil' THEN 'Plussmobil'
            WHEN {name} = 'happybytes' THEN 'Happybytes'
            WHEN {name} = 'unifon' THEN 'Unifon'
            ELSE {other}
        END
    """


def provider_total_price_group(column: str = "fusnavn") -> str:
    name = f"lower({column})"
    return f"""
        CASE
            WHEN contains({name}, 'telenor') THEN 'Telenor'
            WHEN contains({name}, 'telia') THEN 'Telia'
            WHEN contains({name}, 'lyse') OR contains({name}, 'ice') THEN 'Ice'
            ELSE 'Øvrige'
        END
    """


def linear_projection(rows: list[dict], periods_ahead: int = 3) -> list[dict]:
    grouped: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in rows:
        grouped[(row["metric"], row["tilbyder"])].append(row)

    projection = []
    for (metric, provider), provider_rows in grouped.items():
        provider_rows = sorted(provider_rows, key=lambda item: item["ar"])
        if len(provider_rows) < 2:
            continue
        xs = [float(row["ar"]) for row in provider_rows]
        ys = [float(row["value"]) for row in provider_rows]
        x_mean = sum(xs) / len(xs)
        y_mean = sum(ys) / len(ys)
        denominator = sum((x - x_mean) ** 2 for x in xs)
        slope = (
            sum((x - x_mean) * (y - y_mean) for x, y in zip(xs, ys)) / denominator
            if denominator
            else 0.0
        )
        latest_year = int(max(xs))
        latest_value = ys[-1]
        for year in range(latest_year, latest_year + periods_ahead + 1):
            projection.append(
                {
                    "metric": metric,
                    "ar": year,
                    "tilbyder": provider,
                    "value": latest_value + slope * (year - latest_year),
                    "series": "Lineær trend",
                }
            )
    return projection


def default_access_owner(provider: str) -> str:
    name = provider.lower()
    if "lyse" in name or "ice" in name:
        return "Lyse Tele (Ice)"
    if name in {"telia norge", "chili mobil", "fjordkraft mobil"}:
        return "Telia"
    return "Telenor"


def display_provider(provider: str) -> str:
    names = {
        "telenor norge": "Telenor",
        "telia norge": "Telia",
        "lyse tele": "Lyse Tele (Ice)",
        "fjordkraft mobil": "Fjordkraft",
        "lycamobile norway ltd": "Lycamobile",
        "xplora mobile": "Xplora",
    }
    return names.get(provider.lower(), provider[:1].upper() + provider[1:])


def normalize_alias(value: object) -> str:
    return " ".join(str(value or "").strip().lower().split())


def provider_alias_map() -> dict[str, str]:
    aliases: dict[str, str] = {}
    for row in run_duckdb(
        """
        SELECT DISTINCT
            lower(fusnavn) AS fusnavn,
            lower(levnavn) AS levnavn,
            lower(grnavn) AS grnavn
        FROM read_parquet('data/mobil.parquet')
        WHERE dk = 'Mobiltelefoni'
            AND hg = 'Abonnement'
            AND n1 IN ('Fakturert', 'Kontantkort')
            AND n2 = 'Ingen'
            AND tp = 'Sum'
            AND sk = 'Sluttbruker'
            AND delar = 'Helår'
        ORDER BY fusnavn, levnavn;
        """
    ):
        provider = normalize_alias(row["fusnavn"])
        for key in ["fusnavn", "levnavn", "grnavn"]:
            alias = normalize_alias(row[key])
            if alias:
                aliases.setdefault(alias, provider)

    aliases.update(
        {
            "ice norge": "lyse tele",
            "lycamobile": "lycamobile norway ltd",
        }
    )
    return aliases


def build_wholesale_assignment_template(provider_subscriptions: list[dict]) -> dict[str, dict[str, str]]:
    providers_by_year: dict[int, set[str]] = defaultdict(set)
    for row in provider_subscriptions:
        providers_by_year[int(row["ar"])].add(normalize_alias(row["provider"]))

    template: dict[str, dict[str, str]] = {
        str(year): {
            provider: default_access_owner(provider)
            for provider in sorted(providers)
        }
        for year, providers in sorted(providers_by_year.items())
    }
    if not ACCESS_BUYER_WORKBOOK.exists():
        return template

    aliases = provider_alias_map()
    workbook = load_workbook(ACCESS_BUYER_WORKBOOK, data_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        owner, year_value, provider_list = row[:3]
        owner = str(owner or "").strip()
        if owner not in WHOLESALE_OWNERS or not year_value or not provider_list:
            continue
        year = int(year_value)
        year_key = str(year)
        if year_key not in template:
            continue
        for raw_provider in str(provider_list).split(","):
            alias = normalize_alias(raw_provider)
            provider = aliases.get(alias, alias)
            if provider in providers_by_year[year]:
                template[year_key][provider] = owner

    return template


def write_csv(filename: str, rows: list[dict], columns: list[str]) -> dict:
    path = EXPORT_DIR / filename
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    return {"csvFilename": filename, "csvPath": f"assets/exports/{filename}"}


def sheet_title(value: str, used: set[str]) -> str:
    forbidden = "[]:*?/\\"
    title = "".join("-" if char in forbidden else char for char in value).strip()[:31] or "Ark"
    candidate = title
    counter = 2
    while candidate in used:
        suffix = f" {counter}"
        candidate = f"{title[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def year_values(rows: list[dict]) -> list[int]:
    return sorted({int(row["ar"]) for row in rows if row.get("ar") is not None})


def as_excel_value(value: object, kind: str) -> object:
    if value is None:
        return None
    if isinstance(value, str):
        try:
            number = float(value)
        except ValueError:
            return value
    else:
        number = float(value) if isinstance(value, (int, float)) else value
    if kind == "percent" and isinstance(number, float):
        return number / 100 if abs(number) > 1 else number
    if kind == "number" and isinstance(number, float) and number.is_integer():
        return int(number)
    return number


def number_format(kind: str) -> str:
    if kind == "percent":
        return "0.0%"
    if kind == "number":
        return "# ##0"
    if kind == "currency":
        return "# ##0.0"
    if kind == "decimal":
        return "0.0000"
    if kind == "hhi":
        return "0.00"
    return "General"


def rows_by_year(rows: list[dict], row_key: str, value_key: str) -> dict[tuple[object, int], object]:
    return {
        (row[row_key], int(row["ar"])): row.get(value_key)
        for row in rows
        if row.get(row_key) is not None and row.get("ar") is not None
    }


def infer_row_order(sections: list[dict], row_key: str, preferred: list[str] | None = None) -> list[object]:
    keys = {
        row[row_key]
        for section in sections
        for row in section["rows"]
        if row.get(row_key) is not None
    }
    if preferred:
        ordered = [key for key in preferred if key in keys]
        ordered.extend(sorted(keys - set(ordered), key=lambda value: str(value).lower()))
        return ordered
    first_section = sections[0]
    years = first_section.get("years") or year_values(first_section["rows"])
    latest = max(years) if years else None
    latest_values = rows_by_year(first_section["rows"], row_key, first_section["value_key"])

    def sort_key(key: object) -> tuple[int, float, str]:
        value = latest_values.get((key, latest)) if latest is not None else None
        return (0 if isinstance(value, (int, float)) else 1, -float(value or 0), str(key).lower())

    return sorted(keys, key=sort_key)


def write_wide_sheet(workbook: Workbook, config: dict, used_titles: set[str]) -> None:
    sheet = workbook.create_sheet(sheet_title(config["name"], used_titles))
    sheet.sheet_view.showGridLines = False

    row_key = config["row_key"]
    row_label = config.get("row_label", row_key)
    sections = config["sections"]
    row_order = infer_row_order(sections, row_key, config.get("row_order"))

    navy = "0B2B66"
    history_fill = "E9F2DF"
    trend_fill = "DDEBF7"
    neutral_fill = "EEF3F8"
    section_fills = [history_fill, trend_fill, "FCE4D6", "E2F0D9"]
    thin = Side(style="thin", color="FFFFFF")
    grid = Side(style="thin", color="D7E0E8")
    medium = Side(style="medium", color="7F8DA5")

    total_columns = 1 + sum(len(section.get("years") or year_values(section["rows"])) for section in sections)
    sheet["A1"] = config["title"]
    sheet["A1"].font = Font(color=navy, bold=True, size=14)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(total_columns, 2))
    sheet.append([])

    column = 2
    section_metadata = []
    for index, section in enumerate(sections):
        years = section.get("years") or year_values(section["rows"])
        if not years:
            continue
        start = column
        end = column + len(years) - 1
        fill = section.get("fill") or section_fills[index % len(section_fills)]
        if start != end:
            sheet.merge_cells(start_row=3, start_column=start, end_row=3, end_column=end)
        cell = sheet.cell(3, start)
        cell.value = section.get("label", "")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(start, end + 1):
            header_cell = sheet.cell(3, col)
            header_cell.fill = PatternFill("solid", fgColor=fill)
            header_cell.border = Border(
                top=medium,
                bottom=grid,
                left=medium if col == start and start > 2 else thin,
                right=thin,
            )
        section_metadata.append((section, years, start, end, fill))
        column = end + 1

    sheet.cell(4, 1).value = row_label
    for col in range(1, total_columns + 1):
        cell = sheet.cell(4, col)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=grid)

    column = 2
    for _, years, start, _, _ in section_metadata:
        for year in years:
            cell = sheet.cell(4, column)
            cell.value = str(year)
            cell.border = Border(bottom=grid, left=medium if column == start and start > 2 else thin)
            column += 1

    for row_index, row_name in enumerate(row_order, start=5):
        label_cell = sheet.cell(row_index, 1)
        label_cell.value = row_name
        label_cell.fill = PatternFill("solid", fgColor=neutral_fill)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.border = Border(bottom=grid)
        column = 2
        for section, years, start, _, fill in section_metadata:
            lookup = rows_by_year(section["rows"], row_key, section["value_key"])
            for year in years:
                cell = sheet.cell(row_index, column)
                cell.value = as_excel_value(lookup.get((row_name, year)), section["kind"])
                cell.number_format = number_format(section["kind"])
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = Border(
                    bottom=grid,
                    left=medium if column == start and start > 2 else thin,
                )
                column += 1

    sheet.freeze_panes = "A5"
    last_column = get_column_letter(max(total_columns, 1))
    sheet.auto_filter.ref = f"A4:{last_column}{max(sheet.max_row, 4)}"
    sheet.column_dimensions["A"].width = max(17, min(32, max(len(str(value)) for value in row_order + [row_label]) + 2))
    for col in range(2, total_columns + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 13


def default_xlsx_sheets(title: str, rows: list[dict], columns: list[tuple[str, str, str]]) -> list[dict]:
    column_kinds = {key: kind for key, _, kind in columns}
    value_key = "value" if any(key == "value" for key, _, _ in columns) else columns[-1][0]
    row_key = next(
        (
            key
            for key, _, kind in columns
            if kind == "text" and key not in {"period", "series", "metric", "segment"}
        ),
        "metric",
    )
    return [
        {
            "name": title,
            "title": title,
            "row_key": row_key,
            "row_label": row_key,
            "sections": [
                {
                    "label": "Helår",
                    "rows": rows,
                    "value_key": value_key,
                    "kind": column_kinds.get(value_key, "number"),
                }
            ],
        }
    ]


def write_xlsx(
    filename: str,
    title: str,
    rows: list[dict],
    columns: list[tuple[str, str, str]],
    sheets: list[dict] | None = None,
) -> dict:
    path = EXPORT_DIR / filename
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles: set[str] = set()
    for sheet_config in sheets or default_xlsx_sheets(title, rows, columns):
        write_wide_sheet(workbook, sheet_config, used_titles)

    workbook.save(path)
    return {"xlsxFilename": filename, "xlsxPath": f"assets/exports/{filename}"}


def write_export(
    key: str,
    basename: str,
    title: str,
    rows: list[dict],
    columns: list[tuple[str, str, str]],
    xlsx_sheets: list[dict] | None = None,
) -> dict:
    csv_info = write_csv(
        f"{basename}.csv",
        rows,
        [column_key for column_key, _, _ in columns],
    )
    xlsx_info = write_xlsx(f"{basename}.xlsx", title, rows, columns, xlsx_sheets)
    return {**csv_info, **xlsx_info, "label": title}


def matching(rows: list[dict], **criteria: object) -> list[dict]:
    return [
        row
        for row in rows
        if all(row.get(key) == value for key, value in criteria.items())
    ]


def value_sheet(
    name: str,
    title: str,
    rows: list[dict],
    row_key: str,
    row_label: str,
    value_key: str = "value",
    kind: str = "percent",
    row_order: list[str] | None = None,
    section_label: str = "Helår",
) -> dict:
    return {
        "name": name,
        "title": title,
        "row_key": row_key,
        "row_label": row_label,
        "row_order": row_order,
        "sections": [
            {
                "label": section_label,
                "rows": rows,
                "value_key": value_key,
                "kind": kind,
            }
        ],
    }


def metric_value_sheets(
    title: str,
    rows: list[dict],
    row_key: str,
    row_label: str,
    metrics: list[str],
    kind: str = "percent",
    row_order: list[str] | None = None,
) -> list[dict]:
    return [
        value_sheet(
            metric,
            f"{title} - {metric.lower()}",
            matching(rows, metric=metric),
            row_key,
            row_label,
            kind=kind,
            row_order=row_order,
        )
        for metric in metrics
    ]


def segment_metric_sheets(
    title: str,
    rows: list[dict],
    row_key: str,
    row_label: str,
    row_order: list[str],
) -> list[dict]:
    sheets = []
    for metric in ["Abonnement", "Omsetning"]:
        for segment in ["Privat", "Bedrift"]:
            sheets.append(
                value_sheet(
                    f"{metric} {segment}",
                    f"{title} - {metric.lower()} {segment.lower()}",
                    matching(rows, metric=metric, segment=segment),
                    row_key,
                    row_label,
                    row_order=row_order,
                )
            )
    return sheets


def projection_sheets(
    history_rows: list[dict],
    trend_rows: list[dict],
    metrics: list[str],
    latest_year: int,
    row_order: list[str],
) -> list[dict]:
    sheets = []
    for metric in metrics:
        sheets.append(
            {
                "name": metric,
                "title": f"Lineær trend basert på {metric.lower()}",
                "row_key": "tilbyder",
                "row_label": "tilbyder",
                "row_order": row_order,
                "sections": [
                    {
                        "label": "Historikk",
                        "rows": matching(history_rows, metric=metric),
                        "years": list(range(min(year_values(history_rows)), latest_year + 1)),
                        "value_key": "value",
                        "kind": "percent",
                        "fill": "E9F2DF",
                    },
                    {
                        "label": "Lineær trend",
                        "rows": [
                            row
                            for row in matching(trend_rows, metric=metric)
                            if int(row["ar"]) > latest_year
                        ],
                        "years": list(range(latest_year + 1, latest_year + 4)),
                        "value_key": "value",
                        "kind": "percent",
                        "fill": "DDEBF7",
                    },
                ],
            }
        )
    return sheets


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    (DIST / "data").mkdir(parents=True, exist_ok=True)

    metadata = run_duckdb(
        """
        SELECT
            count(*) AS row_count,
            min(ar) AS first_year,
            max(ar) AS latest_year,
            count(DISTINCT fusnavn) AS provider_count,
            count(DISTINCT levnavn) AS operator_count
        FROM read_parquet('data/mobil.parquet');
        """
    )[0]

    market_share = run_duckdb(
        f"""
        WITH base AS (
            SELECT
                ar,
                'Abonnement' AS metric,
                {provider_group()} AS tilbyder,
                sum(svar) AS absolute
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Abonnement'
                AND n1 IN ('Fakturert', 'Kontantkort')
                AND n2 = 'Ingen'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, tilbyder

            UNION ALL

            SELECT
                ar,
                'Omsetning' AS metric,
                {provider_group()} AS tilbyder,
                sum(svar) AS absolute
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Inntekter'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, tilbyder
        )
        SELECT
            metric,
            ar,
            tilbyder,
            absolute,
            100.0 * absolute / sum(absolute) OVER (PARTITION BY metric, ar) AS value
        FROM base
        ORDER BY metric, ar, value DESC;
        """
    )

    projection = linear_projection(market_share)

    segment_share = run_duckdb(
        f"""
        WITH base AS (
            SELECT
                ar,
                ms AS segment,
                'Abonnement' AS metric,
                {provider_group()} AS tilbyder,
                sum(svar) AS absolute
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Abonnement'
                AND ms IN ('Privat', 'Bedrift')
                AND n1 IN ('Fakturert', 'Kontantkort')
                AND n2 = 'Ingen'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, segment, tilbyder

            UNION ALL

            SELECT
                ar,
                ms AS segment,
                'Omsetning' AS metric,
                {provider_group()} AS tilbyder,
                sum(svar) AS absolute
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Inntekter'
                AND ms IN ('Privat', 'Bedrift')
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, segment, tilbyder
        )
        SELECT
            metric,
            segment,
            ar,
            tilbyder,
            absolute,
            100.0 * absolute / sum(absolute) OVER (PARTITION BY metric, segment, ar) AS value
        FROM base
        ORDER BY metric, segment, ar, value DESC;
        """
    )

    private_challengers = run_duckdb(
        """
        WITH years AS (
            SELECT DISTINCT ar
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg IN ('Abonnement', 'Inntekter')
                AND ms = 'Privat'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
        ),
        providers(tilbyder) AS (
            VALUES
                ('Fjordkraft'),
                ('Chili mobil'),
                ('Lycamobile'),
                ('Xplora'),
                ('Happybytes'),
                ('Plussmobil')
        ),
        metrics(metric, hg) AS (
            VALUES ('Abonnement', 'Abonnement'), ('Omsetning', 'Inntekter')
        ),
        rows AS (
            SELECT
                ar,
                CASE WHEN hg = 'Abonnement' THEN 'Abonnement' ELSE 'Omsetning' END AS metric,
                CASE
                    WHEN lower(fusnavn) = 'fjordkraft mobil' THEN 'Fjordkraft'
                    WHEN lower(fusnavn) = 'chili mobil' THEN 'Chili mobil'
                    WHEN lower(fusnavn) = 'lycamobile norway ltd' THEN 'Lycamobile'
                    WHEN lower(fusnavn) = 'xplora mobile' THEN 'Xplora'
                    WHEN lower(fusnavn) = 'happybytes' THEN 'Happybytes'
                    WHEN lower(fusnavn) = 'plussmobil' THEN 'Plussmobil'
                    ELSE NULL
                END AS tilbyder,
                svar,
                sum(svar) OVER (PARTITION BY ar, hg) AS total
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg IN ('Abonnement', 'Inntekter')
                AND ms = 'Privat'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
                AND (hg = 'Inntekter' OR (n1 IN ('Fakturert', 'Kontantkort') AND n2 = 'Ingen'))
        ),
        agg AS (
            SELECT
                ar,
                metric,
                tilbyder,
                sum(svar) AS absolute,
                max(total) AS total
            FROM rows
            WHERE tilbyder IS NOT NULL
            GROUP BY ar, metric, tilbyder
        )
        SELECT
            m.metric,
            y.ar,
            p.tilbyder,
            coalesce(a.absolute, 0) AS absolute,
            coalesce(100.0 * a.absolute / a.total, 0) AS value
        FROM years y
        CROSS JOIN metrics m
        CROSS JOIN providers p
        LEFT JOIN agg a ON a.ar = y.ar AND a.metric = m.metric AND a.tilbyder = p.tilbyder
        ORDER BY m.metric, y.ar, p.tilbyder;
        """
    )

    business_challengers = run_duckdb(
        """
        WITH years AS (
            SELECT DISTINCT ar
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg IN ('Abonnement', 'Inntekter')
                AND ms = 'Bedrift'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
        ),
        providers(tilbyder) AS (
            VALUES ('Unifon'), ('Nortel'), ('Saga mobil'), ('SMB mobil')
        ),
        metrics(metric, hg) AS (
            VALUES ('Abonnement', 'Abonnement'), ('Omsetning', 'Inntekter')
        ),
        rows AS (
            SELECT
                ar,
                CASE WHEN hg = 'Abonnement' THEN 'Abonnement' ELSE 'Omsetning' END AS metric,
                CASE
                    WHEN lower(levnavn) = 'unifon' THEN 'Unifon'
                    WHEN lower(levnavn) = 'nortel' THEN 'Nortel'
                    WHEN lower(levnavn) = 'saga mobil' THEN 'Saga mobil'
                    WHEN lower(levnavn) = 'smb mobil' THEN 'SMB mobil'
                    ELSE NULL
                END AS tilbyder,
                svar,
                sum(svar) OVER (PARTITION BY ar, hg) AS total
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg IN ('Abonnement', 'Inntekter')
                AND ms = 'Bedrift'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
                AND (hg = 'Inntekter' OR (n1 IN ('Fakturert', 'Kontantkort') AND n2 = 'Ingen'))
        ),
        agg AS (
            SELECT
                ar,
                metric,
                tilbyder,
                sum(svar) AS absolute,
                max(total) AS total
            FROM rows
            WHERE tilbyder IS NOT NULL
            GROUP BY ar, metric, tilbyder
        )
        SELECT
            m.metric,
            y.ar,
            p.tilbyder,
            coalesce(a.absolute, 0) AS absolute,
            coalesce(100.0 * a.absolute / a.total, 0) AS value
        FROM years y
        CROSS JOIN metrics m
        CROSS JOIN providers p
        LEFT JOIN agg a ON a.ar = y.ar AND a.metric = m.metric AND a.tilbyder = p.tilbyder
        ORDER BY m.metric, y.ar, p.tilbyder;
        """
    )

    arpu_segment = run_duckdb(
        """
        WITH abonnement_snapshot AS (
            SELECT ar, ms AS segment, sum(svar) AS abonnement
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Abonnement'
                AND ms IN ('Privat', 'Bedrift')
                AND n1 IN ('Fakturert', 'Kontantkort')
                AND n2 = 'Ingen'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, segment
        ),
        abonnement AS (
            SELECT
                current.ar,
                current.segment,
                current.abonnement AS abonnement_slutt_ar,
                previous.abonnement AS abonnement_forrige_ar,
                (current.abonnement + coalesce(previous.abonnement, current.abonnement)) / 2.0
                    AS abonnement
            FROM abonnement_snapshot current
            LEFT JOIN abonnement_snapshot previous
                ON previous.segment = current.segment
                AND previous.ar = current.ar - 1
        ),
        omsetning AS (
            SELECT ar, ms AS segment, sum(svar) AS omsetning
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Inntekter'
                AND ms IN ('Privat', 'Bedrift')
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, segment
        )
        SELECT
            o.ar,
            o.segment,
            o.omsetning,
            a.abonnement,
            a.abonnement_forrige_ar,
            a.abonnement_slutt_ar,
            o.omsetning::DOUBLE * 1000 / a.abonnement / 12 AS value
        FROM omsetning o
        JOIN abonnement a USING (ar, segment)
        ORDER BY segment, ar;
        """
    )

    arpu_provider = run_duckdb(
        f"""
        WITH abonnement_snapshot AS (
            SELECT
                ar,
                ms AS segment,
                {provider_price_group()} AS tilbyder,
                sum(svar) AS abonnement
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Abonnement'
                AND ms IN ('Privat', 'Bedrift')
                AND n1 IN ('Fakturert', 'Kontantkort')
                AND n2 = 'Ingen'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, segment, tilbyder
        ),
        abonnement AS (
            SELECT
                current.ar,
                current.segment,
                current.tilbyder,
                current.abonnement AS abonnement_slutt_ar,
                previous.abonnement AS abonnement_forrige_ar,
                (current.abonnement + coalesce(previous.abonnement, current.abonnement)) / 2.0
                    AS abonnement
            FROM abonnement_snapshot current
            LEFT JOIN abonnement_snapshot previous
                ON previous.segment = current.segment
                AND previous.tilbyder = current.tilbyder
                AND previous.ar = current.ar - 1
        ),
        omsetning AS (
            SELECT
                ar,
                ms AS segment,
                {provider_price_group()} AS tilbyder,
                sum(svar) AS omsetning
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Inntekter'
                AND ms IN ('Privat', 'Bedrift')
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, segment, tilbyder
        )
        SELECT
            o.ar,
            o.segment,
            o.tilbyder,
            o.omsetning,
            a.abonnement,
            a.abonnement_forrige_ar,
            a.abonnement_slutt_ar,
            o.omsetning::DOUBLE * 1000 / a.abonnement / 12 AS value
        FROM omsetning o
        JOIN abonnement a USING (ar, segment, tilbyder)
        WHERE o.tilbyder IS NOT NULL
        ORDER BY segment, tilbyder, ar;
        """
    )

    nok_per_gb_total = run_duckdb(
        f"""
        WITH inntekter AS (
            SELECT
                ar,
                {provider_total_price_group()} AS tilbyder,
                sum(svar) AS inntekter
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Inntekter'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, tilbyder
        ),
        trafikk AS (
            SELECT
                ar,
                {provider_total_price_group()} AS tilbyder,
                sum(svar) AS datatrafikk_gb
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Trafikk'
                AND n1 = 'Data'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, tilbyder
        )
        SELECT
            i.ar,
            i.tilbyder,
            i.inntekter,
            t.datatrafikk_gb,
            i.inntekter::DOUBLE * 1000 / t.datatrafikk_gb AS value
        FROM inntekter i
        JOIN trafikk t USING (ar, tilbyder)
        ORDER BY tilbyder, ar;
        """
    )

    nok_per_gb_providers = run_duckdb(
        f"""
        WITH inntekter AS (
            SELECT
                ar,
                {provider_price_group()} AS tilbyder,
                sum(svar) AS inntekter
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Inntekter'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, tilbyder
        ),
        trafikk AS (
            SELECT
                ar,
                {provider_price_group()} AS tilbyder,
                sum(svar) AS datatrafikk_gb
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Trafikk'
                AND n1 = 'Data'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, tilbyder
        )
        SELECT
            i.ar,
            i.tilbyder,
            i.inntekter,
            t.datatrafikk_gb,
            i.inntekter::DOUBLE * 1000 / t.datatrafikk_gb AS value
        FROM inntekter i
        JOIN trafikk t USING (ar, tilbyder)
        WHERE i.tilbyder IS NOT NULL
        ORDER BY tilbyder, ar;
        """
    )

    provider_subscriptions = run_duckdb(
        """
        SELECT
            ar,
            delar AS period,
            lower(fusnavn) AS provider,
            sum(svar) AS abonnement
        FROM read_parquet('data/mobil.parquet')
        WHERE dk = 'Mobiltelefoni'
            AND hg = 'Abonnement'
            AND n1 IN ('Fakturert', 'Kontantkort')
            AND n2 = 'Ingen'
            AND tp = 'Sum'
            AND sk = 'Sluttbruker'
            AND delar = 'Helår'
        GROUP BY ar, period, provider
        ORDER BY ar, abonnement DESC;
        """
    )
    provider_meta = [
        {
            "provider": row["provider"],
            "label": display_provider(row["provider"]),
            "defaultOwner": default_access_owner(row["provider"]),
        }
        for row in run_duckdb(
            """
            SELECT provider
            FROM (
                SELECT lower(fusnavn) AS provider, sum(svar) AS abonnement
                FROM read_parquet('data/mobil.parquet')
                WHERE dk = 'Mobiltelefoni'
                    AND hg = 'Abonnement'
                    AND n1 IN ('Fakturert', 'Kontantkort')
                    AND n2 = 'Ingen'
                    AND tp = 'Sum'
                    AND sk = 'Sluttbruker'
                    AND delar = 'Helår'
                GROUP BY provider
            )
            WHERE abonnement > 0
            ORDER BY abonnement DESC;
            """
        )
    ]

    concentration = run_duckdb(
        """
        WITH base AS (
            SELECT
                ar,
                'Abonnement' AS metric,
                fusnavn,
                sum(svar) AS absolute
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Abonnement'
                AND n1 IN ('Fakturert', 'Kontantkort')
                AND n2 = 'Ingen'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, fusnavn

            UNION ALL

            SELECT
                ar,
                'Omsetning' AS metric,
                fusnavn,
                sum(svar) AS absolute
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Inntekter'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, fusnavn
        ),
        shares AS (
            SELECT
                ar,
                metric,
                fusnavn,
                absolute::DOUBLE / sum(absolute) OVER (PARTITION BY metric, ar) AS share
            FROM base
        ),
        ranked AS (
            SELECT
                *,
                row_number() OVER (PARTITION BY metric, ar ORDER BY share DESC) AS rank
            FROM shares
        )
        SELECT
            metric,
            ar,
            100.0 * sum(CASE WHEN rank <= 2 THEN share ELSE 0 END) AS cr2,
            sum(share * share) AS hhi
        FROM ranked
        GROUP BY metric, ar
        ORDER BY metric, ar;
        """
    )

    totals = run_duckdb(
        """
        WITH abonnement AS (
            SELECT
                ar,
                delar AS period,
                'Abonnement' AS metric,
                sum(svar) AS value
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Abonnement'
                AND n1 IN ('Fakturert', 'Kontantkort')
                AND n2 = 'Ingen'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, period
        ),
        inntekter AS (
            SELECT
                ar,
                delar AS period,
                'Inntekter' AS metric,
                sum(svar) AS value
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Inntekter'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, period
        )
        SELECT * FROM abonnement
        UNION ALL
        SELECT * FROM inntekter
        ORDER BY metric, period, ar;
        """
    )

    provider_share_trend = run_duckdb(
        f"""
        WITH base AS (
            SELECT
                ar,
                delar AS period,
                'Abonnement' AS metric,
                {provider_group()} AS tilbyder,
                sum(svar) AS absolute
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Abonnement'
                AND n1 IN ('Fakturert', 'Kontantkort')
                AND n2 = 'Ingen'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, period, tilbyder

            UNION ALL

            SELECT
                ar,
                delar AS period,
                'Omsetning' AS metric,
                {provider_group()} AS tilbyder,
                sum(svar) AS absolute
            FROM read_parquet('data/mobil.parquet')
            WHERE dk = 'Mobiltelefoni'
                AND hg = 'Inntekter'
                AND tp = 'Sum'
                AND sk = 'Sluttbruker'
                AND delar = 'Helår'
            GROUP BY ar, period, tilbyder
        )
        SELECT
            metric,
            period,
            ar,
            tilbyder,
            absolute,
            100.0 * absolute / sum(absolute) OVER (PARTITION BY metric, period, ar) AS value
        FROM base
        ORDER BY metric, period, ar, value DESC;
        """
    )

    wholesale_assignment_template = build_wholesale_assignment_template(provider_subscriptions)
    wholesale_buckets: dict[tuple[str, int, str], dict[str, float]] = defaultdict(
        lambda: {"abonnement": 0.0}
    )
    period_totals: dict[tuple[str, int], float] = defaultdict(float)
    for row in provider_subscriptions:
        key = (row["period"], int(row["ar"]))
        owner = wholesale_assignment_template.get(str(row["ar"]), {}).get(
            row["provider"],
            default_access_owner(row["provider"]),
        )
        value = float(row["abonnement"])
        wholesale_buckets[(row["period"], int(row["ar"]), owner)]["abonnement"] += value
        period_totals[key] += value

    wholesale_default = []
    for (period, year, owner), values in sorted(wholesale_buckets.items()):
        total = period_totals[(period, year)]
        wholesale_default.append(
            {
                "period": period,
                "ar": year,
                "grossist": owner,
                "abonnement": round(values["abonnement"]),
                "value": values["abonnement"] * 100 / total if total else 0,
            }
        )

    grossist_concentration_default = []
    for (period, year), total in sorted(period_totals.items()):
        shares = [
            row["value"] / 100
            for row in wholesale_default
            if row["period"] == period and row["ar"] == year
        ]
        grossist_concentration_default.append(
            {
                "period": period,
                "ar": year,
                "cr2": sum(sorted(shares, reverse=True)[:2]) * 100,
                "hhi": sum(share * share for share in shares),
            }
        )

    group_order = ["Telenor", "Telia", "Lyse Tele (Ice)", "Øvrige"]
    private_order = ["Fjordkraft", "Chili mobil", "Lycamobile", "Xplora", "Happybytes", "Plussmobil"]
    business_order = ["Unifon", "Nortel", "Saga mobil", "SMB mobil"]
    price_order = [
        "Telenor",
        "Telia",
        "Ice",
        "Chili mobil",
        "Happybytes",
        "Fjordkraft",
        "Plussmobil",
        "Unifon",
    ]

    exports = {
        "market-share": write_export(
            "market-share",
            "market-share",
            "Markedsandeler",
            market_share,
            [
                ("metric", "Grunnlag", "text"),
                ("ar", "År", "number"),
                ("tilbyder", "Tilbyder", "text"),
                ("absolute", "Verdi", "number"),
                ("value", "Markedsandel (%)", "percent"),
            ],
            xlsx_sheets=metric_value_sheets(
                "Markedsandeler",
                market_share,
                "tilbyder",
                "tilbyder",
                ["Abonnement", "Omsetning"],
                row_order=group_order,
            ),
        ),
        "projection": write_export(
            "projection",
            "market-share-projection",
            "Lineær trend i markedsandeler",
            projection,
            [
                ("metric", "Grunnlag", "text"),
                ("ar", "År", "number"),
                ("tilbyder", "Tilbyder", "text"),
                ("series", "Serie", "text"),
                ("value", "Markedsandel (%)", "percent"),
            ],
            xlsx_sheets=projection_sheets(
                market_share,
                projection,
                ["Abonnement", "Omsetning"],
                int(metadata["latest_year"]),
                group_order,
            ),
        ),
        "segment-share": write_export(
            "segment-share",
            "segment-share",
            "Markedsandeler per segment",
            segment_share,
            [
                ("metric", "Grunnlag", "text"),
                ("segment", "Segment", "text"),
                ("ar", "År", "number"),
                ("tilbyder", "Tilbyder", "text"),
                ("absolute", "Verdi", "number"),
                ("value", "Markedsandel (%)", "percent"),
            ],
            xlsx_sheets=segment_metric_sheets(
                "Markedsandeler per segment",
                segment_share,
                "tilbyder",
                "tilbyder",
                group_order,
            ),
        ),
        "private-challengers": write_export(
            "private-challengers",
            "private-challengers",
            "Privatmarkedet øvrige tilbydere",
            private_challengers,
            [
                ("metric", "Grunnlag", "text"),
                ("ar", "År", "number"),
                ("tilbyder", "Tilbyder", "text"),
                ("absolute", "Verdi", "number"),
                ("value", "Markedsandel (%)", "percent"),
            ],
            xlsx_sheets=metric_value_sheets(
                "Privatmarkedet øvrige tilbydere",
                private_challengers,
                "tilbyder",
                "tilbyder",
                ["Abonnement", "Omsetning"],
                row_order=private_order,
            ),
        ),
        "business-challengers": write_export(
            "business-challengers",
            "business-challengers",
            "Bedriftsmarkedet øvrige tilbydere",
            business_challengers,
            [
                ("metric", "Grunnlag", "text"),
                ("ar", "År", "number"),
                ("tilbyder", "Tilbyder", "text"),
                ("absolute", "Verdi", "number"),
                ("value", "Markedsandel (%)", "percent"),
            ],
            xlsx_sheets=metric_value_sheets(
                "Bedriftsmarkedet øvrige tilbydere",
                business_challengers,
                "tilbyder",
                "tilbyder",
                ["Abonnement", "Omsetning"],
                row_order=business_order,
            ),
        ),
        "arpu-segment": write_export(
            "arpu-segment",
            "arpu-segment",
            "Omsetning per kunde per segment",
            arpu_segment,
            [
                ("ar", "År", "number"),
                ("segment", "Segment", "text"),
                ("omsetning", "Omsetning", "number"),
                ("abonnement_forrige_ar", "Abonnement forrige helår", "number"),
                ("abonnement_slutt_ar", "Abonnement inneværende helår", "number"),
                ("abonnement", "Abonnement snitt", "number"),
                ("value", "NOK per måned", "currency"),
            ],
            xlsx_sheets=[
                value_sheet(
                    "ARPU",
                    "Omsetning per kunde per segment",
                    arpu_segment,
                    "segment",
                    "segment",
                    kind="currency",
                    row_order=["Bedrift", "Privat"],
                )
            ],
        ),
        "arpu-provider": write_export(
            "arpu-provider",
            "arpu-provider",
            "Omsetning per kunde per tilbyder",
            arpu_provider,
            [
                ("ar", "År", "number"),
                ("segment", "Segment", "text"),
                ("tilbyder", "Tilbyder", "text"),
                ("omsetning", "Omsetning", "number"),
                ("abonnement_forrige_ar", "Abonnement forrige helår", "number"),
                ("abonnement_slutt_ar", "Abonnement inneværende helår", "number"),
                ("abonnement", "Abonnement snitt", "number"),
                ("value", "NOK per måned", "currency"),
            ],
            xlsx_sheets=[
                value_sheet(
                    "Privat",
                    "Omsetning per kunde per tilbyder - privat",
                    matching(arpu_provider, segment="Privat"),
                    "tilbyder",
                    "tilbyder",
                    kind="currency",
                    row_order=price_order,
                ),
                value_sheet(
                    "Bedrift",
                    "Omsetning per kunde per tilbyder - bedrift",
                    matching(arpu_provider, segment="Bedrift"),
                    "tilbyder",
                    "tilbyder",
                    kind="currency",
                    row_order=price_order,
                ),
            ],
        ),
        "nok-per-gb-total": write_export(
            "nok-per-gb-total",
            "nok-per-gb-total",
            "Omsetning per GB totalt",
            nok_per_gb_total,
            [
                ("ar", "År", "number"),
                ("tilbyder", "Tilbyder", "text"),
                ("inntekter", "Inntekter", "number"),
                ("datatrafikk_gb", "Datatrafikk GB", "number"),
                ("value", "NOK per GB", "currency"),
            ],
            xlsx_sheets=[
                value_sheet(
                    "NOK per GB",
                    "Omsetning per GB totalt",
                    nok_per_gb_total,
                    "tilbyder",
                    "tilbyder",
                    kind="currency",
                    row_order=["Telenor", "Telia", "Ice", "Øvrige"],
                )
            ],
        ),
        "nok-per-gb-providers": write_export(
            "nok-per-gb-providers",
            "nok-per-gb-providers",
            "Omsetning per GB tilbydere",
            nok_per_gb_providers,
            [
                ("ar", "År", "number"),
                ("tilbyder", "Tilbyder", "text"),
                ("inntekter", "Inntekter", "number"),
                ("datatrafikk_gb", "Datatrafikk GB", "number"),
                ("value", "NOK per GB", "currency"),
            ],
            xlsx_sheets=[
                value_sheet(
                    "NOK per GB",
                    "Omsetning per GB tilbydere",
                    nok_per_gb_providers,
                    "tilbyder",
                    "tilbyder",
                    kind="currency",
                    row_order=price_order,
                )
            ],
        ),
        "wholesale": write_export(
            "wholesale",
            "wholesale-market-shares",
            "Wholesale market shares",
            wholesale_default,
            [
                ("period", "Periode", "text"),
                ("ar", "År", "number"),
                ("grossist", "Grossist", "text"),
                ("abonnement", "Abonnement", "number"),
                ("value", "Andel (%)", "percent"),
            ],
            xlsx_sheets=[
                {
                    "name": "Grossistmarked",
                    "title": "Wholesale market shares",
                    "row_key": "grossist",
                    "row_label": "grossist",
                    "row_order": WHOLESALE_OWNERS,
                    "sections": [
                        {
                            "label": "Andel",
                            "rows": wholesale_default,
                            "value_key": "value",
                            "kind": "percent",
                            "fill": "E9F2DF",
                        },
                        {
                            "label": "Abonnement",
                            "rows": wholesale_default,
                            "value_key": "abonnement",
                            "kind": "number",
                            "fill": "DDEBF7",
                        },
                    ],
                }
            ],
        ),
        "concentration": write_export(
            "concentration",
            "market-concentration",
            "Markedskonsentrasjon",
            concentration,
            [
                ("metric", "Grunnlag", "text"),
                ("ar", "År", "number"),
                ("cr2", "CR2 (%)", "percent"),
                ("hhi", "HHI", "decimal"),
            ],
            xlsx_sheets=[
                {
                    "name": "Konsentrasjon",
                    "title": "Markedskonsentrasjon",
                    "row_key": "metric",
                    "row_label": "grunnlag",
                    "row_order": ["Abonnement", "Omsetning"],
                    "sections": [
                        {
                            "label": "CR2",
                            "rows": concentration,
                            "value_key": "cr2",
                            "kind": "percent",
                            "fill": "E9F2DF",
                        },
                        {
                            "label": "HHI",
                            "rows": concentration,
                            "value_key": "hhi",
                            "kind": "hhi",
                            "fill": "DDEBF7",
                        },
                    ],
                }
            ],
        ),
        "totals": write_export(
            "totals",
            "totals",
            "Totale abonnement og inntekter",
            totals,
            [
                ("metric", "Grunnlag", "text"),
                ("period", "Periode", "text"),
                ("ar", "År", "number"),
                ("value", "Verdi", "number"),
            ],
            xlsx_sheets=[
                value_sheet(
                    "Totaler",
                    "Totale abonnement og inntekter",
                    totals,
                    "metric",
                    "grunnlag",
                    value_key="value",
                    kind="number",
                    row_order=["Abonnement", "Inntekter"],
                )
            ],
        ),
        "provider-share-trend": write_export(
            "provider-share-trend",
            "provider-share-trend",
            "Tilbyderandeler abonnement og inntekt",
            provider_share_trend,
            [
                ("metric", "Grunnlag", "text"),
                ("period", "Periode", "text"),
                ("ar", "År", "number"),
                ("tilbyder", "Tilbyder", "text"),
                ("absolute", "Verdi", "number"),
                ("value", "Andel (%)", "percent"),
            ],
            xlsx_sheets=metric_value_sheets(
                "Tilbyderandeler abonnement og inntekt",
                provider_share_trend,
                "tilbyder",
                "tilbyder",
                ["Abonnement", "Omsetning"],
                row_order=group_order,
            ),
        ),
    }

    payload = {
        "metadata": {
            **metadata,
            "source": "data/mobil.parquet",
            "built_from": "DuckDB CLI",
        },
        "order": {
            "groups": ["Telenor", "Telia", "Lyse Tele (Ice)", "Øvrige"],
            "privateChallengers": [
                "Fjordkraft",
                "Chili mobil",
                "Lycamobile",
                "Xplora",
                "Happybytes",
                "Plussmobil",
            ],
            "businessChallengers": ["Unifon", "Nortel", "Saga mobil", "SMB mobil"],
            "priceProviders": [
                "Telenor",
                "Telia",
                "Ice",
                "Chili mobil",
                "Happybytes",
                "Fjordkraft",
                "Plussmobil",
                "Unifon",
            ],
            "wholesaleOwners": WHOLESALE_OWNERS,
        },
        "exports": exports,
        "marketShare": market_share,
        "projection": projection,
        "segmentShare": segment_share,
        "privateChallengers": private_challengers,
        "businessChallengers": business_challengers,
        "arpuSegment": arpu_segment,
        "arpuProvider": arpu_provider,
        "nokPerGbTotal": nok_per_gb_total,
        "nokPerGbProviders": nok_per_gb_providers,
        "providerSubscriptions": provider_subscriptions,
        "providers": provider_meta,
        "wholesaleAssignmentTemplate": wholesale_assignment_template,
        "wholesaleDefault": wholesale_default,
        "grossistConcentrationDefault": grossist_concentration_default,
        "concentration": concentration,
        "totals": totals,
        "providerShareTrend": provider_share_trend,
    }

    (DATA_DIR / "app-data.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    shutil.copy2(SOURCE_PARQUET, DIST / "data" / "mobil.parquet")
    print(
        "Built app data:",
        len(market_share),
        "market-share rows,",
        len(segment_share),
        "segment rows, latest year",
        metadata["latest_year"],
    )


if __name__ == "__main__":
    main()
