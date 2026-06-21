# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "marimo>=0.23.3",
#   "matplotlib>=3.8.0",
#   "openpyxl>=3.1.0",
#   "pyarrow>=16.0.0",
#   "polars>=1.0.0",
# ]
# ///

import marimo

__generated_with = "0.23.8"
app = marimo.App(width="full")


@app.cell
def _(df, mo, pl, table_excel_download):
    def _concentration(_metric, _label):
        _metric_data = (
            df.filter(
                (pl.col("dk") == "Mobiltelefoni")
                & (pl.col("hg") == _metric)
                & (pl.col("tp") == "Sum")
                & (pl.col("sk") == "Sluttbruker")
                & (pl.col("delar") == "Helår")
                & (
                    (_metric == "Inntekter")
                    | (
                        (pl.col("n1").is_in(["Fakturert", "Kontantkort"]))
                        & (pl.col("n2") == "Ingen")
                    )
                )
            )
            .group_by("ar", "fusnavn")
            .agg(pl.col("svar").sum().alias("verdi"))
            .with_columns(
                markedsandel=pl.col("verdi")
                / pl.col("verdi").sum().over("ar")
            )
            .collect()
        )
        return (
            _metric_data.group_by("ar")
            .agg(
                pl.col("markedsandel").sort(descending=True).head(2).sum().alias("CR2"),
                (pl.col("markedsandel") ** 2).sum().alias("HHI"),
            )
            .with_columns(
                pl.lit(_label).alias("grunnlag"),
                pl.col("ar").cast(pl.Utf8),
            )
            .select("grunnlag", "ar", "CR2", "HHI")
        )

    concentration = (
        pl.concat(
            [
                _concentration("Abonnement", "Basert på abonnement (sluttbrukermarked)"),
                _concentration("Inntekter", "Basert på omsetning (sluttbrukermarked)"),
            ]
        )
        .sort("grunnlag", "ar")
    )
    _display = concentration.with_columns(
        (pl.col("CR2") * 100).round(1).alias("CR2 (%)"),
        pl.col("HHI").round(4).alias("HHI"),
    ).select("grunnlag", "ar", "CR2 (%)", "HHI")
    _export = table_excel_download(
        concentration,
        "figur-12-markedskonsentrasjon.xlsx",
        "Figur 12 - Markedskonsentrasjon",
        "Konsentrasjon",
        _percent_columns=["CR2"],
    )
    mo.vstack(
        [
            mo.hstack(
                [
                    mo.Html('<div class="figure-heading-title">CR2 og HHI</div>'),
                    _export,
                ],
                justify="start",
                gap=0.35,
            ),
            mo.ui.table(_display, page_size=14),
        ],
        gap=0.5,
    )
    return


@app.cell
def _():
    from base64 import b64encode
    from io import BytesIO
    from pathlib import Path

    import matplotlib.pyplot as plt
    import marimo as mo
    import polars as pl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    try:
        from pyodide.http import pyfetch
    except ImportError:
        pyfetch = None
    return (
        Alignment,
        Border,
        BytesIO,
        Font,
        Path,
        PatternFill,
        Side,
        Workbook,
        b64encode,
        load_workbook,
        mo,
        pl,
        plt,
        pyfetch,
    )


@app.cell
async def _(Path, pl, pyfetch):
    if pyfetch is None:
        parquet_path = Path(__file__).parent / "data" / "mobil.parquet"
    else:
        response = await pyfetch("../data/mobil.parquet")
        parquet_path = Path("/tmp/mobil.parquet")
        parquet_path.write_bytes(await response.bytes())

    df = pl.scan_parquet(parquet_path)
    return (df,)


@app.cell
def _(
    Alignment,
    Border,
    BytesIO,
    Font,
    PatternFill,
    Side,
    Workbook,
    b64encode,
    mo,
    pl,
):
    def excel_download(_data, _filename, _title, _sheet_name="Data"):
        _is_trend_export = "serie" in _data.columns
        if _is_trend_export:
            _index_columns = [
                _column for _column in ["ms", "tilbyder"] if _column in _data.columns
            ]
            _history_years = sorted(
                _data.filter(pl.col("serie") == "Historikk")["ar"].unique().to_list()
            )
            _last_history_year = max(_history_years)
            _trend_years = sorted(
                _data.filter(
                    (pl.col("serie") == "Lineær trend")
                    & (pl.col("ar") > _last_history_year)
                )["ar"].unique().to_list()
            )
            _export_data = (
                _data.filter(
                    (pl.col("serie") == "Historikk")
                    | (
                        (pl.col("serie") == "Lineær trend")
                        & (pl.col("ar") > _last_history_year)
                    )
                )
                .with_columns(
                    (pl.col("markedsandel") / 100).alias("markedsandel"),
                    pl.col("ar").cast(pl.Utf8),
                )
                .pivot(
                    values="markedsandel",
                    index=_index_columns,
                    on="ar",
                    aggregate_function="first",
                )
                .select(
                    _index_columns
                    + [str(_year) for _year in _history_years + _trend_years]
                )
                .sort(_index_columns)
            )
            _years = _history_years + _trend_years
        else:
            _index_columns = [
                _column for _column in ["ms", "tilbyder"] if _column in _data.columns
            ]
            _years = sorted(_data["ar"].unique().to_list())
            _export_data = (
                _data.with_columns(
                    (pl.col("markedsandel") / 100).alias("markedsandel"),
                    pl.col("ar").cast(pl.Utf8),
                )
                .pivot(
                    values="markedsandel",
                    index=_index_columns,
                    on="ar",
                    aggregate_function="first",
                )
                .select(_index_columns + [str(_year) for _year in _years])
                .sort(_index_columns)
            )
        _headers = _export_data.columns
        _header_row = 4 if _is_trend_export else 3
        _first_data_row = _header_row + 1
        _workbook = Workbook()
        _sheet = _workbook.active
        _sheet.title = _sheet_name[:31]

        _sheet["A1"] = _title
        _sheet["A1"].font = Font(bold=True, size=14, color="0B2B66")
        _sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=max(1, len(_headers)),
        )

        _sheet.append([])
        if _is_trend_export:
            _sheet.append([""] * len(_headers))
            _history_start = len(_index_columns) + 1
            _history_end = _history_start + len(_history_years) - 1
            _trend_start = _history_end + 1
            _trend_end = _trend_start + len(_trend_years) - 1
            if _history_years:
                _sheet.merge_cells(
                    start_row=3,
                    start_column=_history_start,
                    end_row=3,
                    end_column=_history_end,
                )
                _sheet.cell(row=3, column=_history_start).value = "Historikk"
            if _trend_years:
                _sheet.merge_cells(
                    start_row=3,
                    start_column=_trend_start,
                    end_row=3,
                    end_column=_trend_end,
                )
                _sheet.cell(row=3, column=_trend_start).value = "Lineær trend"
        _sheet.append(_headers)
        for _row in _export_data.iter_rows():
            _sheet.append(list(_row))

        _header_fill = PatternFill("solid", fgColor="0B2B66")
        _header_font = Font(bold=True, color="FFFFFF")
        _thin = Side(style="thin", color="D9D9D9")
        _medium_blue = Side(style="medium", color="0B2B66")
        _border = Border(bottom=_thin)

        if _is_trend_export:
            _history_fill = PatternFill("solid", fgColor="E9F2DF")
            _trend_fill = PatternFill("solid", fgColor="DDEBF7")
            for _cell in _sheet[3]:
                _cell.fill = _history_fill
                _cell.font = Font(bold=True, color="000000")
                _cell.alignment = Alignment(horizontal="center")
                _cell.border = Border(top=_medium_blue, bottom=_thin)
            for _column_index in range(_trend_start, _trend_end + 1):
                _cell = _sheet.cell(row=3, column=_column_index)
                _cell.fill = _trend_fill
                _cell.border = Border(top=_medium_blue, bottom=_thin)
            for _row_index in range(3, _sheet.max_row + 1):
                _left_cell = _sheet.cell(row=_row_index, column=_trend_start)
                _left_cell.border = Border(
                    left=_medium_blue,
                    right=_left_cell.border.right,
                    top=_left_cell.border.top,
                    bottom=_left_cell.border.bottom,
                )

        for _cell in _sheet[_header_row]:
            _cell.fill = _header_fill
            _cell.font = _header_font
            _cell.alignment = Alignment(horizontal="center")
            if _is_trend_export and _cell.column >= _trend_start:
                _cell.fill = PatternFill("solid", fgColor="1F4E79")

        for _row in _sheet.iter_rows(
            min_row=_first_data_row,
            max_row=_sheet.max_row,
            max_col=_sheet.max_column,
        ):
            for _cell in _row:
                _cell.border = _border
                if _cell.column_letter == "A":
                    _cell.alignment = Alignment(horizontal="left")
                if _headers[_cell.column - 1] in [str(_year) for _year in _years]:
                    _cell.number_format = "0.0%"

        _table_ref = f"A{_header_row}:{_sheet.cell(_sheet.max_row, _sheet.max_column).coordinate}"
        for _row_index in range(_first_data_row, _sheet.max_row + 1):
            if (_row_index - _first_data_row) % 2 == 0:
                for _cell in _sheet[_row_index]:
                    _cell.fill = PatternFill("solid", fgColor="EEF3F8")
            if _is_trend_export:
                for _column_index in range(len(_index_columns) + 1, _history_end + 1):
                    _sheet.cell(row=_row_index, column=_column_index).fill = PatternFill(
                        "solid",
                        fgColor="F3F8EC",
                    )
                for _column_index in range(_trend_start, _trend_end + 1):
                    _sheet.cell(row=_row_index, column=_column_index).fill = PatternFill(
                        "solid",
                        fgColor="EAF3FB",
                    )
                _left_cell = _sheet.cell(row=_row_index, column=_trend_start)
                _left_cell.border = Border(
                    left=_medium_blue,
                    right=_left_cell.border.right,
                    top=_left_cell.border.top,
                    bottom=_left_cell.border.bottom,
                )
        _sheet.freeze_panes = f"A{_first_data_row}"
        _sheet.auto_filter.ref = _table_ref

        for _column_index, _header in enumerate(_headers, start=1):
            _max_length = len(str(_header))
            for _row_index in range(_first_data_row, _sheet.max_row + 1):
                _value = _sheet.cell(row=_row_index, column=_column_index).value
                _max_length = max(
                    _max_length,
                    len(str(_value)) if _value is not None else 0,
            )
            _sheet.column_dimensions[
                _sheet.cell(row=_header_row, column=_column_index).column_letter
            ].width = min(
                max(_max_length + 2, 12),
                28,
            )

        _buffer = BytesIO()
        _workbook.save(_buffer)
        _payload = b64encode(_buffer.getvalue()).decode("ascii")

        return mo.Html(
            f"""
            <a
                class="excel-download"
                href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{_payload}"
                download="{_filename}"
                aria-label="Last ned Excel"
                title="Last ned Excel"
            >
                <svg viewBox="0 0 24 24" aria-hidden="true">
                    <path d="M12 3v11"></path>
                    <path d="M7.5 9.5 12 14l4.5-4.5"></path>
                    <path d="M5 18h14"></path>
                </svg>
            </a>
            """
        )

    return (excel_download,)


@app.cell
def _(
    Alignment,
    Border,
    BytesIO,
    Font,
    PatternFill,
    Side,
    Workbook,
    b64encode,
    mo,
    pl,
):
    def table_excel_download(_data, _filename, _title, _sheet_name="Data", _percent_columns=None):
        _percent_columns = set(_percent_columns or [])
        _export_data = _data.collect() if isinstance(_data, pl.LazyFrame) else _data
        _headers = _export_data.columns
        _workbook = Workbook()
        _sheet = _workbook.active
        _sheet.title = _sheet_name[:31]

        _sheet["A1"] = _title
        _sheet["A1"].font = Font(bold=True, size=14, color="0B2B66")
        _sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=max(1, len(_headers)),
        )
        _sheet.append([])
        _sheet.append(_headers)
        for _row in _export_data.iter_rows():
            _sheet.append(list(_row))

        _header_fill = PatternFill("solid", fgColor="0B2B66")
        _header_font = Font(bold=True, color="FFFFFF")
        _thin = Side(style="thin", color="D9D9D9")
        _border = Border(bottom=_thin)
        for _cell in _sheet[3]:
            _cell.fill = _header_fill
            _cell.font = _header_font
            _cell.alignment = Alignment(horizontal="center")

        for _row in _sheet.iter_rows(min_row=4, max_row=_sheet.max_row):
            for _cell in _row:
                _cell.border = _border
                _header = _headers[_cell.column - 1]
                if _header in _percent_columns:
                    _cell.number_format = "0.0%"
                elif isinstance(_cell.value, float):
                    _cell.number_format = "0.0"
                if _cell.column == 1:
                    _cell.alignment = Alignment(horizontal="left")

        _sheet.freeze_panes = "A4"
        _sheet.auto_filter.ref = f"A3:{_sheet.cell(_sheet.max_row, _sheet.max_column).coordinate}"
        for _column_index, _header in enumerate(_headers, start=1):
            _max_length = len(str(_header))
            for _row_index in range(4, _sheet.max_row + 1):
                _value = _sheet.cell(row=_row_index, column=_column_index).value
                _max_length = max(_max_length, len(str(_value)) if _value is not None else 0)
            _sheet.column_dimensions[
                _sheet.cell(row=3, column=_column_index).column_letter
            ].width = min(max(_max_length + 2, 12), 42)

        _buffer = BytesIO()
        _workbook.save(_buffer)
        _payload = b64encode(_buffer.getvalue()).decode("ascii")
        return mo.Html(
            f"""
            <a
                class="excel-download"
                href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{_payload}"
                download="{_filename}"
                aria-label="Last ned Excel"
                title="Last ned Excel"
            >
                <svg viewBox="0 0 24 24" aria-hidden="true">
                    <path d="M12 3v11"></path>
                    <path d="M7.5 9.5 12 14l4.5-4.5"></path>
                    <path d="M5 18h14"></path>
                </svg>
            </a>
            """
        )

    return (table_excel_download,)


@app.cell
def _(mo):
    mo.Html(
        """
        <style>
            .figure-heading-title {
                align-items: center;
                display: inline-flex;
                font-size: 1.25rem;
                font-weight: 700;
                line-height: 30px;
            }

            .excel-download {
                align-items: center !important;
                border-radius: 4px !important;
                color: #5f6f82 !important;
                display: inline-flex !important;
                height: 22px !important;
                justify-content: center !important;
                margin: 0 !important;
                text-decoration: none !important;
                transform: translateY(1px);
                width: 22px !important;
            }

            .excel-download:hover {
                background: #f4f7fa !important;
                color: #0b2b66 !important;
            }

            .excel-download svg {
                color: currentColor !important;
                fill: none !important;
                height: 14px !important;
                stroke: currentColor !important;
                stroke-linecap: round !important;
                stroke-linejoin: round !important;
                stroke-width: 1.8 !important;
                width: 14px !important;
            }
        </style>
        <div style="margin-bottom: 20px;">
            <div style="font-size: 2.2rem; font-weight: 700; color: #0b2b66;">
                Mobilanalyse marked
            </div>
        </div>
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ## 1 - Utvikling i markedsandeler
    """)
    return


@app.cell
def _(df, pl):
    _provider_name = pl.col("fusnavn").str.to_lowercase()
    _provider_group = (
        pl.when(_provider_name.str.contains("telenor"))
        .then(pl.lit("Telenor"))
        .when(_provider_name.str.contains("telia"))
        .then(pl.lit("Telia"))
        .when(
            _provider_name.is_in(["ice communication norge", "lyse tele"])
            | _provider_name.str.contains("lyse")
            | _provider_name.str.contains("ice")
        )
        .then(pl.lit("Lyse Tele (Ice)"))
        .otherwise(pl.lit("Øvrige"))
    )

    market_share_abonnement = (
        df.filter(
            (pl.col("dk") == "Mobiltelefoni")
            & (pl.col("hg") == "Abonnement")
            & (pl.col("n1").is_in(["Fakturert", "Kontantkort"]))
            & (pl.col("n2") == "Ingen")
            & (pl.col("tp") == "Sum")
            & (pl.col("sk") == "Sluttbruker")
            & (pl.col("delar") == "Helår")
        )
        .with_columns(tilbyder=_provider_group)
        .group_by("ar", "tilbyder")
        .agg(pl.col("svar").sum().alias("abonnement"))
        .with_columns(
            markedsandel=pl.col("abonnement")
            * 100
            / pl.col("abonnement").sum().over("ar")
        )
        .select("ar", "tilbyder", "markedsandel")
        .sort("ar", "tilbyder")
        .collect()
    )
    return (market_share_abonnement,)


@app.cell
def _(df, pl):
    _provider_name = pl.col("fusnavn").str.to_lowercase()
    _provider_group = (
        pl.when(_provider_name.str.contains("telenor"))
        .then(pl.lit("Telenor"))
        .when(_provider_name.str.contains("telia"))
        .then(pl.lit("Telia"))
        .when(
            _provider_name.is_in(["ice communication norge", "lyse tele"])
            | _provider_name.str.contains("lyse")
            | _provider_name.str.contains("ice")
        )
        .then(pl.lit("Lyse Tele (Ice)"))
        .otherwise(pl.lit("Øvrige"))
    )

    market_share_omsetning = (
        df.filter(
            (pl.col("dk") == "Mobiltelefoni")
            & (pl.col("hg") == "Inntekter")
            & (pl.col("tp") == "Sum")
            & (pl.col("sk") == "Sluttbruker")
            & (pl.col("delar") == "Helår")
        )
        .with_columns(tilbyder=_provider_group)
        .group_by("ar", "tilbyder")
        .agg(pl.col("svar").sum().alias("omsetning"))
        .with_columns(
            markedsandel=pl.col("omsetning")
            * 100
            / pl.col("omsetning").sum().over("ar")
        )
        .select("ar", "tilbyder", "markedsandel")
        .sort("ar", "tilbyder")
        .collect()
    )
    return (market_share_omsetning,)


@app.cell
def _(
    excel_download,
    market_share_abonnement,
    market_share_omsetning,
    mo,
    plt,
):
    _colors = {
        "Telenor": "#00B0F0",
        "Telia": "#9900CC",
        "Lyse Tele (Ice)": "#ffc000",
        "Øvrige": "#00b050",
    }
    _order = ["Telenor", "Telia", "Lyse Tele (Ice)", "Øvrige"]

    def _percent(_value):
        return f"{_value:.1f} %".replace(".", ",")

    def _plot_market_share(_data, _upper_y):
        _fig, _ax = plt.subplots(figsize=(7.2, 4.8), dpi=120)
        for _provider in _order:
            _provider_data = _data.filter(_data["tilbyder"] == _provider).sort("ar")
            if _provider_data.is_empty():
                continue

            _xs = _provider_data["ar"].to_list()
            _ys = _provider_data["markedsandel"].to_list()
            _ax.plot(
                _xs,
                _ys,
                color=_colors[_provider],
                linewidth=3.0,
                solid_capstyle="round",
                label=_provider,
            )

            if len(_ys) >= 2:
                _ax.annotate(
                    _percent(_ys[-2]),
                    xy=(_xs[-2], _ys[-2]),
                    xytext=(10, 6),
                    textcoords="offset points",
                    color="#404040",
                    fontsize=9,
                )
            _ax.annotate(
                _percent(_ys[-1]),
                xy=(_xs[-1], _ys[-1]),
                xytext=(10, 6),
                textcoords="offset points",
                color="#404040",
                fontsize=9,
            )

        _ax.set_ylim(0, _upper_y)
        _ax.set_yticks(range(0, _upper_y + 1, 10))
        _ax.set_yticklabels([_percent(_value) for _value in range(0, _upper_y + 1, 10)])
        _ax.set_xticks(sorted(_data["ar"].unique().to_list()))
        _ax.grid(axis="y", color="#d9d9d9", linewidth=0.8)
        _ax.grid(axis="x", visible=False)
        _ax.spines[["top", "right", "left"]].set_visible(False)
        _ax.spines["bottom"].set_color("#d9d9d9")
        _ax.tick_params(axis="both", colors="#555555", length=0)
        _ax.legend(
            loc="lower center",
            bbox_to_anchor=(0.5, -0.18),
            ncol=4,
            frameon=False,
            fontsize=10,
        )
        _fig.tight_layout()
        return _fig

    _abonnement_fig = _plot_market_share(market_share_abonnement, 60)
    _omsetning_fig = _plot_market_share(market_share_omsetning, 60)
    _abonnement_export = excel_download(
        market_share_abonnement,
        "figur-1-abonnement.xlsx",
        "Figur 1 - Markedsandeler basert på abonnement",
        "Abonnement",
    )
    _omsetning_export = excel_download(
        market_share_omsetning,
        "figur-1-omsetning.xlsx",
        "Figur 1 - Markedsandeler basert på omsetning",
        "Omsetning",
    )

    mo.vstack(
        [
            mo.hstack(
                [
                    mo.vstack(
                        [
                            mo.hstack(
                                [
                                    mo.Html(
                                        '<div class="figure-heading-title">Basert på abonnement</div>'
                                    ),
                                    _abonnement_export,
                                ],
                                justify="start",
                                gap=0.35,
                            ),
                            _abonnement_fig,
                        ],
                        gap=0.5,
                    ),
                    mo.vstack(
                        [
                            mo.hstack(
                                [
                                    mo.Html(
                                        '<div class="figure-heading-title">Basert på omsetning</div>'
                                    ),
                                    _omsetning_export,
                                ],
                                justify="start",
                                gap=0.35,
                            ),
                            _omsetning_fig,
                        ],
                        gap=0.5,
                    ),
                ],
                justify="center",
                gap=2,
            ),
        ],
        gap=2,
    )
    return


@app.cell
def _(mo):
    mo.Html(
        """
        <div style="
            width: 100%;
            margin: 46px 0 24px;
            border-top: 3px solid #0b2b66;
        "></div>
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ## 2 - Lineær trend i markedsandeler
    """)
    return


@app.cell
def _(market_share_abonnement, market_share_omsetning, pl):
    def _linear_projection(_data, _periods_ahead=3):
        _rows = []
        for _provider in ["Telenor", "Telia", "Lyse Tele (Ice)", "Øvrige"]:
            _provider_data = _data.filter(pl.col("tilbyder") == _provider).sort("ar")
            if _provider_data.height < 2:
                continue

            _xs = [float(_x) for _x in _provider_data["ar"].to_list()]
            _ys = [float(_y) for _y in _provider_data["markedsandel"].to_list()]
            _x_mean = sum(_xs) / len(_xs)
            _y_mean = sum(_ys) / len(_ys)
            _denominator = sum((_x - _x_mean) ** 2 for _x in _xs)
            _slope = (
                sum((_x - _x_mean) * (_y - _y_mean) for _x, _y in zip(_xs, _ys))
                / _denominator
                if _denominator
                else 0
            )
            _last_year = int(max(_xs))
            _last_value = _ys[-1]

            for _year in range(_last_year, _last_year + _periods_ahead + 1):
                _rows.append(
                    {
                        "ar": _year,
                        "tilbyder": _provider,
                        "markedsandel": _last_value + _slope * (_year - _last_year),
                    }
                )

        return pl.DataFrame(_rows)

    market_share_abonnement_projection = _linear_projection(market_share_abonnement)
    market_share_omsetning_projection = _linear_projection(market_share_omsetning)
    return (
        market_share_abonnement_projection,
        market_share_omsetning_projection,
    )


@app.cell
def _(
    excel_download,
    market_share_abonnement,
    market_share_abonnement_projection,
    market_share_omsetning,
    market_share_omsetning_projection,
    mo,
    pl,
    plt,
):
    _colors = {
        "Telenor": "#00B0F0",
        "Telia": "#9900CC",
        "Lyse Tele (Ice)": "#ffc000",
        "Øvrige": "#00b050",
    }
    _order = ["Telenor", "Telia", "Lyse Tele (Ice)", "Øvrige"]

    def _percent(_value):
        return f"{_value:.1f} %".replace(".", ",")

    def _plot_projection(_actual, _projection, _title, _upper_y):
        _fig, _ax = plt.subplots(figsize=(7.8, 4.8), dpi=120)
        _last_actual_year = int(_actual["ar"].max())

        for _provider in _order:
            _actual_provider = _actual.filter(pl.col("tilbyder") == _provider).sort("ar")
            _projection_provider = _projection.filter(pl.col("tilbyder") == _provider).sort("ar")
            if _actual_provider.is_empty() or _projection_provider.is_empty():
                continue

            _actual_xs = _actual_provider["ar"].to_list()
            _actual_ys = _actual_provider["markedsandel"].to_list()
            _projection_xs = _projection_provider["ar"].to_list()
            _projection_ys = _projection_provider["markedsandel"].to_list()

            _ax.plot(
                _actual_xs,
                _actual_ys,
                color=_colors[_provider],
                linewidth=2.5,
                solid_capstyle="round",
                label=_provider,
            )
            _ax.plot(
                _projection_xs,
                _projection_ys,
                color=_colors[_provider],
                linewidth=1.8,
                linestyle=":",
                label="_nolegend_",
            )

            _latest_actual = _actual_provider.filter(pl.col("ar") == _last_actual_year)
            if not _latest_actual.is_empty():
                _ax.annotate(
                    _percent(_latest_actual["markedsandel"][0]),
                    xy=(_last_actual_year, _latest_actual["markedsandel"][0]),
                    xytext=(8, 7),
                    textcoords="offset points",
                    color="#404040",
                    fontsize=8,
                )

            _ax.annotate(
                _percent(_projection_ys[-1]),
                xy=(_projection_xs[-1], _projection_ys[-1]),
                xytext=(8, 0),
                textcoords="offset points",
                color="#202020",
                fontsize=8,
                bbox={
                    "boxstyle": "square,pad=0.35",
                    "facecolor": "white",
                    "edgecolor": _colors[_provider],
                    "linewidth": 0.8,
                },
            )

        _ax.set_ylim(0, _upper_y)
        _ax.set_yticks(range(0, _upper_y + 1, 10))
        _ax.set_yticklabels([_percent(_value) for _value in range(0, _upper_y + 1, 10)])
        _ax.set_xticks(
            sorted(set(_actual["ar"].unique().to_list() + _projection["ar"].unique().to_list()))
        )
        _ax.grid(axis="y", color="#d9d9d9", linewidth=0.8)
        _ax.grid(axis="x", visible=False)
        _ax.spines[["top", "right", "left"]].set_visible(False)
        _ax.spines["bottom"].set_color("#d9d9d9")
        _ax.tick_params(axis="both", colors="#555555", length=0)
        _ax.legend(
            loc="lower center",
            bbox_to_anchor=(0.5, -0.26),
            ncol=4,
            frameon=False,
            fontsize=8,
        )
        _fig.tight_layout()
        return _fig

    def _threshold_text(_data, _provider, _threshold):
        _provider_data = _data.filter(pl.col("tilbyder") == _provider).sort("ar")
        _xs = [float(_x) for _x in _provider_data["ar"].to_list()]
        _ys = [float(_y) for _y in _provider_data["markedsandel"].to_list()]
        _x_mean = sum(_xs) / len(_xs)
        _y_mean = sum(_ys) / len(_ys)
        _denominator = sum((_x - _x_mean) ** 2 for _x in _xs)
        _slope = (
            sum((_x - _x_mean) * (_y - _y_mean) for _x, _y in zip(_xs, _ys))
            / _denominator
            if _denominator
            else 0
        )
        if _slope <= 0:
            return f"Lyse Tele (Ice) når ikke {_threshold:.0f} % omsetningsandel med lineær trend."
        _last_year = max(_xs)
        _last_value = _ys[-1]
        _year = _last_year + ((_threshold - _last_value) / _slope)
        return f"Lyse Tele (Ice) når {_threshold:.0f} % av omsetningen rundt {int(_year + 0.999)}."

    def _projection_export_data(_actual, _projection):
        _actual_export = _actual.with_columns(
            pl.lit("Historikk").alias("serie"),
            pl.col("ar").cast(pl.Int64),
        )
        _projection_export = _projection.with_columns(
            pl.lit("Lineær trend").alias("serie"),
            pl.col("ar").cast(pl.Int64),
        )
        return (
            pl.concat([_actual_export, _projection_export])
            .select("serie", "ar", "tilbyder", "markedsandel")
            .sort("serie", "ar", "tilbyder")
        )

    _abonnement_projection_fig = _plot_projection(
        market_share_abonnement,
        market_share_abonnement_projection,
        "Abonnement",
        60,
    )
    _omsetning_projection_fig = _plot_projection(
        market_share_omsetning,
        market_share_omsetning_projection,
        "Omsetning",
        60,
    )
    _abonnement_projection_export = excel_download(
        _projection_export_data(market_share_abonnement, market_share_abonnement_projection),
        "figur-2-abonnement-trend.xlsx",
        "Figur 2 - Lineær trend basert på abonnement",
        "Abonnement trend",
    )
    _omsetning_projection_export = excel_download(
        _projection_export_data(market_share_omsetning, market_share_omsetning_projection),
        "figur-2-omsetning-trend.xlsx",
        "Figur 2 - Lineær trend basert på omsetning",
        "Omsetning trend",
    )
    mo.vstack(
        [
            mo.hstack(
                [
                    mo.vstack(
                        [
                            mo.hstack(
                                [
                                    mo.Html(
                                        '<div class="figure-heading-title">Abonnement</div>'
                                    ),
                                    _abonnement_projection_export,
                                ],
                                justify="start",
                                gap=0.35,
                            ),
                            _abonnement_projection_fig,
                        ],
                        gap=0.5,
                    ),
                    mo.vstack(
                        [
                            mo.hstack(
                                [
                                    mo.Html(
                                        '<div class="figure-heading-title">Omsetning</div>'
                                    ),
                                    _omsetning_projection_export,
                                ],
                                justify="start",
                                gap=0.35,
                            ),
                            _omsetning_projection_fig,
                        ],
                        gap=0.5,
                    ),
                ],
                justify="center",
                gap=2,
            ),
        ],
        gap=1,
    )
    return


@app.cell
def _(mo):
    mo.Html(
        """
        <div style="
            width: 100%;
            margin: 46px 0 24px;
            border-top: 3px solid #0b2b66;
        "></div>
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ## 3 - Abonnement fordelt på privat og bedrift
    """)
    return


@app.cell
def _(df, pl):
    _provider_name = pl.col("fusnavn").str.to_lowercase()
    _provider_group = (
        pl.when(_provider_name.str.contains("telenor"))
        .then(pl.lit("Telenor"))
        .when(_provider_name.str.contains("telia"))
        .then(pl.lit("Telia"))
        .when(
            _provider_name.is_in(["ice communication norge", "lyse tele"])
            | _provider_name.str.contains("lyse")
            | _provider_name.str.contains("ice")
        )
        .then(pl.lit("Lyse Tele (Ice)"))
        .otherwise(pl.lit("Øvrige"))
    )

    market_share_abonnement_segment = (
        df.filter(
            (pl.col("dk") == "Mobiltelefoni")
            & (pl.col("hg") == "Abonnement")
            & (pl.col("ms").is_in(["Privat", "Bedrift"]))
            & (pl.col("n1").is_in(["Fakturert", "Kontantkort"]))
            & (pl.col("n2") == "Ingen")
            & (pl.col("tp") == "Sum")
            & (pl.col("sk") == "Sluttbruker")
            & (pl.col("delar") == "Helår")
        )
        .with_columns(tilbyder=_provider_group)
        .group_by("ar", "ms", "tilbyder")
        .agg(pl.col("svar").sum().alias("abonnement"))
        .with_columns(
            markedsandel=pl.col("abonnement")
            * 100
            / pl.col("abonnement").sum().over(["ar", "ms"])
        )
        .select("ar", "ms", "tilbyder", "markedsandel")
        .sort("ms", "ar", "tilbyder")
        .collect()
    )
    return (market_share_abonnement_segment,)


@app.cell
def _(excel_download, market_share_abonnement_segment, mo, pl, plt):
    _colors = {
        "Telenor": "#00B0F0",
        "Telia": "#9900CC",
        "Lyse Tele (Ice)": "#ffc000",
        "Øvrige": "#00b050",
    }
    _order = ["Telenor", "Telia", "Lyse Tele (Ice)", "Øvrige"]

    def _percent(_value):
        return f"{_value:.1f} %".replace(".", ",")

    def _plot_segment(_segment, _upper_y=60):
        _data = market_share_abonnement_segment.filter(pl.col("ms") == _segment)
        _fig, _ax = plt.subplots(figsize=(7.2, 4.8), dpi=120)

        for _provider in _order:
            _provider_data = _data.filter(pl.col("tilbyder") == _provider).sort("ar")
            _provider_data = _provider_data.filter(pl.col("markedsandel") > 0)
            if _provider_data.is_empty():
                continue

            _xs = _provider_data["ar"].to_list()
            _ys = _provider_data["markedsandel"].to_list()
            _ax.plot(
                _xs,
                _ys,
                color=_colors[_provider],
                linewidth=3.0,
                solid_capstyle="round",
                label=_provider,
            )

            if len(_ys) >= 2:
                _ax.annotate(
                    _percent(_ys[-2]),
                    xy=(_xs[-2], _ys[-2]),
                    xytext=(8, 7),
                    textcoords="offset points",
                    color="#404040",
                    fontsize=9,
                )
            _ax.annotate(
                _percent(_ys[-1]),
                xy=(_xs[-1], _ys[-1]),
                xytext=(10, 0),
                textcoords="offset points",
                color="#404040",
                fontsize=9,
            )

        _ax.set_ylim(0, _upper_y)
        _ax.set_yticks(range(0, _upper_y + 1, 10))
        _ax.set_yticklabels([_percent(_value) for _value in range(0, _upper_y + 1, 10)])
        _ax.set_xticks(sorted(_data["ar"].unique().to_list()))
        _ax.grid(axis="y", color="#d9d9d9", linewidth=0.8)
        _ax.grid(axis="x", visible=False)
        _ax.spines[["top", "right", "left"]].set_visible(False)
        _ax.spines["bottom"].set_color("#d9d9d9")
        _ax.tick_params(axis="both", colors="#555555", length=0)
        _ax.legend(
            loc="lower center",
            bbox_to_anchor=(0.5, -0.22),
            ncol=4,
            frameon=False,
            fontsize=10,
        )
        _fig.tight_layout()
        return _fig

    _private_fig = _plot_segment("Privat", 50)
    _business_fig = _plot_segment("Bedrift")
    _private_export = excel_download(
        market_share_abonnement_segment.filter(pl.col("ms") == "Privat"),
        "figur-3-privat.xlsx",
        "Figur 3 - Abonnement i privatmarkedet",
        "Privat",
    )
    _business_export = excel_download(
        market_share_abonnement_segment.filter(pl.col("ms") == "Bedrift"),
        "figur-3-bedrift.xlsx",
        "Figur 3 - Abonnement i bedriftsmarkedet",
        "Bedrift",
    )
    mo.vstack(
        [
            mo.hstack(
                [
                    mo.vstack(
                        [
                            mo.hstack(
                                [
                                    mo.Html(
                                        '<div class="figure-heading-title">Privat</div>'
                                    ),
                                    _private_export,
                                ],
                                justify="start",
                                gap=0.35,
                            ),
                            _private_fig,
                        ],
                        gap=0.5,
                    ),
                    mo.vstack(
                        [
                            mo.hstack(
                                [
                                    mo.Html(
                                        '<div class="figure-heading-title">Bedrift</div>'
                                    ),
                                    _business_export,
                                ],
                                justify="start",
                                gap=0.35,
                            ),
                            _business_fig,
                        ],
                        gap=0.5,
                    ),
                ],
                justify="center",
                gap=2,
            ),
        ],
        gap=2,
    )
    return


@app.cell
def _(mo):
    mo.Html(
        """
        <div style="
            width: 100%;
            margin: 46px 0 24px;
            border-top: 3px solid #0b2b66;
        "></div>
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ## 4 - Omsetning fordelt på privat og bedrift
    """)
    return


@app.cell
def _(df, pl):
    _provider_name = pl.col("fusnavn").str.to_lowercase()
    _provider_group = (
        pl.when(_provider_name.str.contains("telenor"))
        .then(pl.lit("Telenor"))
        .when(_provider_name.str.contains("telia"))
        .then(pl.lit("Telia"))
        .when(
            _provider_name.is_in(["ice communication norge", "lyse tele"])
            | _provider_name.str.contains("lyse")
            | _provider_name.str.contains("ice")
        )
        .then(pl.lit("Lyse Tele (Ice)"))
        .otherwise(pl.lit("Øvrige"))
    )

    market_share_omsetning_segment = (
        df.filter(
            (pl.col("dk") == "Mobiltelefoni")
            & (pl.col("hg") == "Inntekter")
            & (pl.col("ms").is_in(["Privat", "Bedrift"]))
            & (pl.col("tp") == "Sum")
            & (pl.col("sk") == "Sluttbruker")
            & (pl.col("delar") == "Helår")
        )
        .with_columns(tilbyder=_provider_group)
        .group_by("ar", "ms", "tilbyder")
        .agg(pl.col("svar").sum().alias("omsetning"))
        .with_columns(
            markedsandel=pl.col("omsetning")
            * 100
            / pl.col("omsetning").sum().over(["ar", "ms"])
        )
        .select("ar", "ms", "tilbyder", "markedsandel")
        .sort("ms", "ar", "tilbyder")
        .collect()
    )
    return (market_share_omsetning_segment,)


@app.cell
def _(excel_download, market_share_omsetning_segment, mo, pl, plt):
    _colors = {
        "Telenor": "#00B0F0",
        "Telia": "#9900CC",
        "Lyse Tele (Ice)": "#ffc000",
        "Øvrige": "#00b050",
    }
    _order = ["Telenor", "Telia", "Lyse Tele (Ice)", "Øvrige"]

    def _percent(_value):
        return f"{_value:.1f} %".replace(".", ",")

    def _plot_segment(_segment, _upper_y=70):
        _data = market_share_omsetning_segment.filter(pl.col("ms") == _segment)
        _fig, _ax = plt.subplots(figsize=(7.2, 4.8), dpi=120)

        for _provider in _order:
            _provider_data = _data.filter(pl.col("tilbyder") == _provider).sort("ar")
            _provider_data = _provider_data.filter(pl.col("markedsandel") > 0)
            if _provider_data.is_empty():
                continue

            _xs = _provider_data["ar"].to_list()
            _ys = _provider_data["markedsandel"].to_list()
            _ax.plot(
                _xs,
                _ys,
                color=_colors[_provider],
                linewidth=3.0,
                solid_capstyle="round",
                label=_provider,
            )

            if len(_ys) >= 2:
                _ax.annotate(
                    _percent(_ys[-2]),
                    xy=(_xs[-2], _ys[-2]),
                    xytext=(8, 7),
                    textcoords="offset points",
                    color="#404040",
                    fontsize=9,
                )
            _ax.annotate(
                _percent(_ys[-1]),
                xy=(_xs[-1], _ys[-1]),
                xytext=(10, 0),
                textcoords="offset points",
                color="#404040",
                fontsize=9,
            )

        _ax.set_ylim(0, _upper_y)
        _ax.set_yticks(range(0, _upper_y + 1, 10))
        _ax.set_yticklabels([_percent(_value) for _value in range(0, _upper_y + 1, 10)])
        _ax.set_xticks(sorted(_data["ar"].unique().to_list()))
        _ax.grid(axis="y", color="#d9d9d9", linewidth=0.8)
        _ax.grid(axis="x", visible=False)
        _ax.spines[["top", "right", "left"]].set_visible(False)
        _ax.spines["bottom"].set_color("#d9d9d9")
        _ax.tick_params(axis="both", colors="#555555", length=0)
        _ax.legend(
            loc="lower center",
            bbox_to_anchor=(0.5, -0.22),
            ncol=4,
            frameon=False,
            fontsize=10,
        )
        _fig.tight_layout()
        return _fig

    _private_fig = _plot_segment("Privat", 60)
    _business_fig = _plot_segment("Bedrift", 70)
    _private_export = excel_download(
        market_share_omsetning_segment.filter(pl.col("ms") == "Privat"),
        "figur-4-privat-omsetning.xlsx",
        "Figur 4 - Omsetning i privatmarkedet",
        "Privat omsetning",
    )
    _business_export = excel_download(
        market_share_omsetning_segment.filter(pl.col("ms") == "Bedrift"),
        "figur-4-bedrift-omsetning.xlsx",
        "Figur 4 - Omsetning i bedriftsmarkedet",
        "Bedrift omsetning",
    )
    mo.vstack(
        [
            mo.hstack(
                [
                    mo.vstack(
                        [
                            mo.hstack(
                                [
                                    mo.Html(
                                        '<div class="figure-heading-title">Privat</div>'
                                    ),
                                    _private_export,
                                ],
                                justify="start",
                                gap=0.35,
                            ),
                            _private_fig,
                        ],
                        gap=0.5,
                    ),
                    mo.vstack(
                        [
                            mo.hstack(
                                [
                                    mo.Html(
                                        '<div class="figure-heading-title">Bedrift</div>'
                                    ),
                                    _business_export,
                                ],
                                justify="start",
                                gap=0.35,
                            ),
                            _business_fig,
                        ],
                        gap=0.5,
                    ),
                ],
                justify="center",
                gap=2,
            ),
        ],
        gap=2,
    )
    return


@app.cell
def _(mo):
    mo.Html(
        """
        <div style="
            width: 100%;
            margin: 46px 0 24px;
            border-top: 3px solid #0b2b66;
        "></div>
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ## 5 - Privatmarkedet: øvrige tilbydere
    """)
    return


@app.cell
def _(df, pl):
    _provider_name = pl.col("fusnavn").str.to_lowercase()
    _private_challenger = (
        pl.when(_provider_name == "fjordkraft mobil")
        .then(pl.lit("Fjordkraft"))
        .when(_provider_name == "chili mobil")
        .then(pl.lit("Chili mobil"))
        .when(_provider_name == "lycamobile norway ltd")
        .then(pl.lit("Lycamobile"))
        .when(_provider_name == "xplora mobile")
        .then(pl.lit("Xplora"))
        .when(_provider_name == "happybytes")
        .then(pl.lit("Happybytes"))
        .when(_provider_name == "plussmobil")
        .then(pl.lit("Plussmobil"))
        .otherwise(None)
    )

    private_challenger_market_share = (
        df.filter(
            (pl.col("dk") == "Mobiltelefoni")
            & (pl.col("hg").is_in(["Abonnement", "Inntekter"]))
            & (pl.col("ms") == "Privat")
            & (pl.col("tp") == "Sum")
            & (pl.col("sk") == "Sluttbruker")
            & (pl.col("delar") == "Helår")
            & (
                (pl.col("hg") == "Inntekter")
                | (
                    (pl.col("n1").is_in(["Fakturert", "Kontantkort"]))
                    & (pl.col("n2") == "Ingen")
                )
            )
        )
        .with_columns(tilbyder=_private_challenger)
        .with_columns(
            total=pl.col("svar").sum().over(["ar", "hg"]),
        )
        .filter(pl.col("tilbyder").is_not_null())
        .group_by("ar", "hg", "tilbyder")
        .agg(
            pl.col("svar").sum().alias("verdi"),
            pl.col("total").first().alias("total"),
        )
        .with_columns(markedsandel=pl.col("verdi") * 100 / pl.col("total"))
        .select("ar", "hg", "tilbyder", "markedsandel")
        .sort("hg", "ar", "tilbyder")
        .collect()
    )
    _years = sorted(private_challenger_market_share["ar"].unique().to_list())
    _metrics = ["Abonnement", "Inntekter"]
    _providers = [
        "Fjordkraft",
        "Chili mobil",
        "Lycamobile",
        "Xplora",
        "Happybytes",
        "Plussmobil",
    ]
    _complete_grid = pl.DataFrame(
        [
            {"ar": _year, "hg": _metric, "tilbyder": _provider}
            for _year in _years
            for _metric in _metrics
            for _provider in _providers
        ]
    )
    private_challenger_market_share = (
        _complete_grid.join(
            private_challenger_market_share,
            on=["ar", "hg", "tilbyder"],
            how="left",
        )
        .with_columns(pl.col("markedsandel").fill_null(0))
        .sort("hg", "ar", "tilbyder")
    )
    return (private_challenger_market_share,)


@app.cell
def _(excel_download, mo, pl, plt, private_challenger_market_share):
    _colors = {
        "Fjordkraft": "#385624",
        "Chili mobil": "#FF0000",
        "Lycamobile": "#002060",
        "Xplora": "#66FF99",
        "Happybytes": "#698ED0",
        "Plussmobil": "#A5A5A5",
    }
    _order = [
        "Fjordkraft",
        "Chili mobil",
        "Lycamobile",
        "Xplora",
        "Happybytes",
        "Plussmobil",
    ]

    def _percent(_value):
        return f"{_value:.1f} %".replace(".", ",")

    def _plot_challengers(_metric, _upper_y=4):
        _data = private_challenger_market_share.filter(pl.col("hg") == _metric)
        _fig, _ax = plt.subplots(figsize=(7.2, 4.8), dpi=120)

        for _provider in _order:
            _provider_data = _data.filter(pl.col("tilbyder") == _provider).sort("ar")
            _provider_data = _provider_data.filter(pl.col("markedsandel") > 0)
            if _provider_data.is_empty():
                continue
            _xs = _provider_data["ar"].to_list()
            _ys = _provider_data["markedsandel"].to_list()
            _ax.plot(
                _xs,
                _ys,
                color=_colors[_provider],
                linewidth=2.6,
                solid_capstyle="round",
                label=_provider,
            )
            if _ys[-1] > 0.05:
                _ax.annotate(
                    _percent(_ys[-1]),
                    xy=(_xs[-1], _ys[-1]),
                    xytext=(8, 0),
                    textcoords="offset points",
                    color="#404040",
                    fontsize=8,
                )

        _ax.set_ylim(0, _upper_y)
        _ax.set_yticks([_value / 2 for _value in range(0, _upper_y * 2 + 1)])
        _ax.set_yticklabels(
            [_percent(_value / 2) for _value in range(0, _upper_y * 2 + 1)]
        )
        _ax.set_xticks(sorted(_data["ar"].unique().to_list()))
        _ax.grid(axis="y", color="#d9d9d9", linewidth=0.8)
        _ax.grid(axis="x", visible=False)
        _ax.spines[["top", "right", "left"]].set_visible(False)
        _ax.spines["bottom"].set_color("#d9d9d9")
        _ax.tick_params(axis="both", colors="#555555", length=0)
        _ax.legend(
            loc="lower center",
            bbox_to_anchor=(0.5, -0.28),
            ncol=3,
            frameon=False,
            fontsize=9,
        )
        _fig.tight_layout()
        return _fig

    _abonnement_fig = _plot_challengers("Abonnement", 4)
    _omsetning_fig = _plot_challengers("Inntekter", 3)
    _abonnement_export = excel_download(
        private_challenger_market_share.filter(pl.col("hg") == "Abonnement"),
        "figur-5-privat-utfordrere-abonnement.xlsx",
        "Figur 5 - Privatmarkedet øvrige tilbydere, abonnement",
        "Abonnement",
    )
    _omsetning_export = excel_download(
        private_challenger_market_share.filter(pl.col("hg") == "Inntekter"),
        "figur-5-privat-utfordrere-omsetning.xlsx",
        "Figur 5 - Privatmarkedet øvrige tilbydere, omsetning",
        "Omsetning",
    )

    mo.hstack(
        [
            mo.vstack(
                [
                    mo.hstack(
                        [
                            mo.Html('<div class="figure-heading-title">Abonnement</div>'),
                            _abonnement_export,
                        ],
                        justify="start",
                        gap=0.35,
                    ),
                    _abonnement_fig,
                ],
                gap=0.5,
            ),
            mo.vstack(
                [
                    mo.hstack(
                        [
                            mo.Html('<div class="figure-heading-title">Omsetning</div>'),
                            _omsetning_export,
                        ],
                        justify="start",
                        gap=0.35,
                    ),
                    _omsetning_fig,
                ],
                gap=0.5,
            ),
        ],
        justify="center",
        gap=2,
    )
    return


@app.cell
def _(mo):
    mo.Html(
        """
        <div style="
            width: 100%;
            margin: 46px 0 24px;
            border-top: 3px solid #0b2b66;
        "></div>
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ## 6 - Bedriftsmarkedet: øvrige tilbydere
    """)
    return


@app.cell
def _(df, pl):
    _provider_name = pl.col("levnavn").str.to_lowercase()
    _business_challenger = (
        pl.when(_provider_name == "unifon")
        .then(pl.lit("Unifon"))
        .when(_provider_name == "nortel")
        .then(pl.lit("Nortel"))
        .when(_provider_name == "saga mobil")
        .then(pl.lit("Saga mobil"))
        .when(_provider_name == "smb mobil")
        .then(pl.lit("SMB mobil"))
        .otherwise(None)
    )

    business_challenger_market_share = (
        df.filter(
            (pl.col("dk") == "Mobiltelefoni")
            & (pl.col("hg").is_in(["Abonnement", "Inntekter"]))
            & (pl.col("ms") == "Bedrift")
            & (pl.col("tp") == "Sum")
            & (pl.col("sk") == "Sluttbruker")
            & (pl.col("delar") == "Helår")
            & (
                (pl.col("hg") == "Inntekter")
                | (
                    (pl.col("n1").is_in(["Fakturert", "Kontantkort"]))
                    & (pl.col("n2") == "Ingen")
                )
            )
        )
        .with_columns(tilbyder=_business_challenger)
        .with_columns(
            total=pl.col("svar").sum().over(["ar", "hg"]),
        )
        .filter(pl.col("tilbyder").is_not_null())
        .group_by("ar", "hg", "tilbyder")
        .agg(
            pl.col("svar").sum().alias("verdi"),
            pl.col("total").first().alias("total"),
        )
        .with_columns(markedsandel=pl.col("verdi") * 100 / pl.col("total"))
        .select("ar", "hg", "tilbyder", "markedsandel")
        .sort("hg", "ar", "tilbyder")
        .collect()
    )
    _years = sorted(business_challenger_market_share["ar"].unique().to_list())
    _metrics = ["Abonnement", "Inntekter"]
    _providers = ["Unifon", "Nortel", "Saga mobil", "SMB mobil"]
    _complete_grid = pl.DataFrame(
        [
            {"ar": _year, "hg": _metric, "tilbyder": _provider}
            for _year in _years
            for _metric in _metrics
            for _provider in _providers
        ]
    )
    business_challenger_market_share = (
        _complete_grid.join(
            business_challenger_market_share,
            on=["ar", "hg", "tilbyder"],
            how="left",
        )
        .with_columns(pl.col("markedsandel").fill_null(0))
        .sort("hg", "ar", "tilbyder")
    )
    return (business_challenger_market_share,)


@app.cell
def _(business_challenger_market_share, excel_download, mo, pl, plt):
    _colors = {
        "Unifon": "#7F7F7F",
        "Nortel": "#64F600",
        "Saga mobil": "#C00000",
        "SMB mobil": "#548235",
    }
    _order = ["Unifon", "Nortel", "Saga mobil", "SMB mobil"]

    def _percent(_value):
        return f"{_value:.1f} %".replace(".", ",")

    def _plot_challengers(_metric, _upper_y=10):
        _data = business_challenger_market_share.filter(pl.col("hg") == _metric)
        _fig, _ax = plt.subplots(figsize=(7.2, 4.8), dpi=120)

        for _provider in _order:
            _provider_data = _data.filter(pl.col("tilbyder") == _provider).sort("ar")
            _provider_data = _provider_data.filter(pl.col("markedsandel") > 0)
            if _provider_data.is_empty():
                continue
            _xs = _provider_data["ar"].to_list()
            _ys = _provider_data["markedsandel"].to_list()
            _ax.plot(
                _xs,
                _ys,
                color=_colors[_provider],
                linewidth=2.8,
                solid_capstyle="round",
                label=_provider,
            )
            if _ys[-1] > 0.05:
                _ax.annotate(
                    _percent(_ys[-1]),
                    xy=(_xs[-1], _ys[-1]),
                    xytext=(8, 0),
                    textcoords="offset points",
                    color="#404040",
                    fontsize=8,
                )

        _ax.set_ylim(0, _upper_y)
        _ax.set_yticks(range(0, _upper_y + 1, 1))
        _ax.set_yticklabels([_percent(_value) for _value in range(0, _upper_y + 1, 1)])
        _ax.set_xticks(sorted(_data["ar"].unique().to_list()))
        _ax.grid(axis="y", color="#d9d9d9", linewidth=0.8)
        _ax.grid(axis="x", visible=False)
        _ax.spines[["top", "right", "left"]].set_visible(False)
        _ax.spines["bottom"].set_color("#d9d9d9")
        _ax.tick_params(axis="both", colors="#555555", length=0)
        _ax.legend(
            loc="lower center",
            bbox_to_anchor=(0.5, -0.22),
            ncol=4,
            frameon=False,
            fontsize=9,
        )
        _fig.tight_layout()
        return _fig

    _abonnement_fig = _plot_challengers("Abonnement", 8)
    _omsetning_fig = _plot_challengers("Inntekter", 10)
    _abonnement_export = excel_download(
        business_challenger_market_share.filter(pl.col("hg") == "Abonnement"),
        "figur-6-bedrift-utfordrere-abonnement.xlsx",
        "Figur 6 - Bedriftsmarkedet øvrige tilbydere, abonnement",
        "Abonnement",
    )
    _omsetning_export = excel_download(
        business_challenger_market_share.filter(pl.col("hg") == "Inntekter"),
        "figur-6-bedrift-utfordrere-omsetning.xlsx",
        "Figur 6 - Bedriftsmarkedet øvrige tilbydere, omsetning",
        "Omsetning",
    )

    mo.hstack(
        [
            mo.vstack(
                [
                    mo.hstack(
                        [
                            mo.Html('<div class="figure-heading-title">Abonnement</div>'),
                            _abonnement_export,
                        ],
                        justify="start",
                        gap=0.35,
                    ),
                    _abonnement_fig,
                ],
                gap=0.5,
            ),
            mo.vstack(
                [
                    mo.hstack(
                        [
                            mo.Html('<div class="figure-heading-title">Omsetning</div>'),
                            _omsetning_export,
                        ],
                        justify="start",
                        gap=0.35,
                    ),
                    _omsetning_fig,
                ],
                gap=0.5,
            ),
        ],
        justify="center",
        gap=2,
    )
    return


@app.cell
def _(mo):
    mo.Html(
        """
        <div style="
            width: 100%;
            margin: 46px 0 24px;
            border-top: 3px solid #0b2b66;
        "></div>
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ## 7 - Oversikt over tilgangskjøpere
    """)
    return


@app.cell
async def _(Path, load_workbook, pl, pyfetch):
    if pyfetch is None:
        _xlsx_path = Path(__file__).parent / "data" / "2025 ekomstat FIGUR - ny format_KAD.xlsx"
    else:
        _response = await pyfetch("../data/2025 ekomstat FIGUR - ny format_KAD.xlsx")
        _xlsx_path = Path("/tmp/ekomstat_figurer.xlsx")
        _xlsx_path.write_bytes(await _response.bytes())

    _workbook = load_workbook(_xlsx_path, read_only=True, data_only=True)
    _sheet = _workbook["Tlgangskjøper oversikt"]
    _rows = list(_sheet.iter_rows(values_only=True))
    _headers = [str(_value).strip() for _value in _rows[0][:3]]
    access_buyers = pl.DataFrame(
        [
            dict(zip(_headers, _row[:3]))
            for _row in _rows[1:]
            if _row[0] is not None and _row[1] is not None and _row[2] is not None
        ]
    ).with_columns(
        pl.col("ar").cast(pl.Int64),
        pl.col("Tilbyder(-e)").str.replace_all(", ", "\n").alias("Tilbydere"),
    )
    return


@app.cell
def _(mo, pl, table_excel_download):
    _summary = pl.DataFrame(
        [
            {
                "Kategori": "MVNO",
                "April 2024 (på vedtakstidspunktet)": "Com4\nHappybytes (hos Telavox)\nLycamobile",
                "November 2025": "Com4\nHappybytes\nLycamobile\nSaga Mobil",
            },
            {
                "Kategori": "Tjenesteleverandører",
                "April 2024 (på vedtakstidspunktet)": (
                    "Chilimobil\nFjordkraft\nUnifon\nIntility (M2M/IoT)\n"
                    "Plussmobil (Telavox)\nPrimafon (Telavox)\n"
                    "Saga Mobil (Telavox)\nXplora Mobil (Telavox)\nSMB Mobil (Svea)"
                ),
                "November 2025": (
                    "Chilimobil\nFjordkraft\nUnifon\nIntility (M2M/IoT)\n"
                    "Plussmobil\nPrimafon\nXplora Mobil\nSMB Mobil\nProject42\n"
                    "B2B Mobil\nBillity (MVNE)\nPingaway\nPhoneIT\nTrøndermobil (i prosess)"
                ),
            },
        ]
    )
    _export = table_excel_download(
        _summary,
        "figur-7-tilgangskjopere.xlsx",
        "Figur 7 - Oversikt over tilgangskjøpere",
        "Tilgangskjøpere",
    )
    mo.vstack(
        [
            mo.hstack(
                [
                    mo.Html('<div class="figure-heading-title">Oversikt over tilgangskjøpere</div>'),
                    _export,
                ],
                justify="start",
                gap=0.35,
            ),
            mo.ui.table(_summary, page_size=4),
        ],
        gap=0.5,
    )
    return


@app.cell
def _(mo):
    mo.Html(
        """
        <div style="
            width: 100%;
            margin: 46px 0 24px;
            border-top: 3px solid #0b2b66;
        "></div>
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ## 8 - Prisutvikling: omsetning per kunde
    """)
    return


@app.cell
def _(df, pl):
    _abonnement = (
        df.filter(
            (pl.col("dk") == "Mobiltelefoni")
            & (pl.col("hg") == "Abonnement")
            & (pl.col("ms").is_in(["Privat", "Bedrift"]))
            & (pl.col("n1").is_in(["Fakturert", "Kontantkort"]))
            & (pl.col("n2") == "Ingen")
            & (pl.col("tp") == "Sum")
            & (pl.col("sk") == "Sluttbruker")
            & (pl.col("delar") == "Helår")
        )
        .group_by("ar", "ms")
        .agg(pl.col("svar").sum().alias("abonnement"))
    )
    _omsetning = (
        df.filter(
            (pl.col("dk") == "Mobiltelefoni")
            & (pl.col("hg") == "Inntekter")
            & (pl.col("ms").is_in(["Privat", "Bedrift"]))
            & (pl.col("tp") == "Sum")
            & (pl.col("sk") == "Sluttbruker")
            & (pl.col("delar") == "Helår")
        )
        .group_by("ar", "ms")
        .agg(pl.col("svar").sum().alias("omsetning"))
    )
    arpu_segment = (
        _omsetning.join(_abonnement, on=["ar", "ms"])
        .with_columns((pl.col("omsetning").cast(pl.Float64) * 1000 / pl.col("abonnement") / 12).alias("arpu"))
        .select("ar", "ms", "arpu")
        .sort("ms", "ar")
        .collect()
    )
    return (arpu_segment,)


@app.cell
def _(arpu_segment, mo, pl, plt, table_excel_download):
    _colors = {"Privat": "#00B0F0", "Bedrift": "#9900CC"}

    def _plot_arpu_segment():
        _fig, _ax = plt.subplots(figsize=(7.8, 4.6), dpi=120)
        for _segment in ["Bedrift", "Privat"]:
            _data = arpu_segment.filter(pl.col("ms") == _segment).sort("ar")
            _ax.plot(
                _data["ar"].to_list(),
                _data["arpu"].to_list(),
                color=_colors[_segment],
                linewidth=3,
                solid_capstyle="round",
                label=f"ARPU {_segment.lower()}markedet",
            )
            _ax.annotate(
                f"{_data['arpu'][-1]:.0f}",
                xy=(_data["ar"][-1], _data["arpu"][-1]),
                xytext=(8, 0),
                textcoords="offset points",
                fontsize=9,
                color="#404040",
            )
        _ax.set_ylabel("NOK per måned")
        _ax.set_xticks(sorted(arpu_segment["ar"].unique().to_list()))
        _ax.grid(axis="y", color="#d9d9d9", linewidth=0.8)
        _ax.spines[["top", "right", "left"]].set_visible(False)
        _ax.spines["bottom"].set_color("#d9d9d9")
        _ax.tick_params(axis="both", colors="#555555", length=0)
        _ax.legend(loc="lower center", bbox_to_anchor=(0.5, -0.22), ncol=2, frameon=False)
        _fig.tight_layout()
        return _fig

    _export = table_excel_download(
        arpu_segment,
        "figur-8-arpu-segment.xlsx",
        "Figur 8 - Omsetning per kunde",
        "ARPU segment",
    )
    mo.vstack(
        [
            mo.hstack(
                [
                    mo.Html('<div class="figure-heading-title">Omsetning per kunde</div>'),
                    _export,
                ],
                justify="start",
                gap=0.35,
            ),
            _plot_arpu_segment(),
        ],
        gap=0.5,
    )
    return


@app.cell
def _(mo):
    mo.Html(
        """
        <div style="
            width: 100%;
            margin: 46px 0 24px;
            border-top: 3px solid #0b2b66;
        "></div>
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ## 9 - Store variasjoner i omsetning per kunde
    """)
    return


@app.cell
def _(df, pl):
    _provider_name = pl.col("fusnavn").str.to_lowercase()
    _provider = (
        pl.when(_provider_name.str.contains("telenor"))
        .then(pl.lit("Telenor"))
        .when(_provider_name.str.contains("telia"))
        .then(pl.lit("Telia"))
        .when(_provider_name.str.contains("lyse") | _provider_name.str.contains("ice"))
        .then(pl.lit("Ice"))
        .when(_provider_name == "fjordkraft mobil")
        .then(pl.lit("Fjordkraft"))
        .when(_provider_name == "chili mobil")
        .then(pl.lit("Chili mobil"))
        .when(_provider_name == "plussmobil")
        .then(pl.lit("Plussmobil"))
        .when(_provider_name == "happybytes")
        .then(pl.lit("Happybytes"))
        .when(_provider_name == "unifon")
        .then(pl.lit("Unifon"))
        .otherwise(None)
    )
    _abonnement = (
        df.filter(
            (pl.col("dk") == "Mobiltelefoni")
            & (pl.col("hg") == "Abonnement")
            & (pl.col("ms").is_in(["Privat", "Bedrift"]))
            & (pl.col("n1").is_in(["Fakturert", "Kontantkort"]))
            & (pl.col("n2") == "Ingen")
            & (pl.col("tp") == "Sum")
            & (pl.col("sk") == "Sluttbruker")
            & (pl.col("delar") == "Helår")
        )
        .with_columns(tilbyder=_provider)
        .filter(pl.col("tilbyder").is_not_null())
        .group_by("ar", "ms", "tilbyder")
        .agg(pl.col("svar").sum().alias("abonnement"))
    )
    _omsetning = (
        df.filter(
            (pl.col("dk") == "Mobiltelefoni")
            & (pl.col("hg") == "Inntekter")
            & (pl.col("ms").is_in(["Privat", "Bedrift"]))
            & (pl.col("tp") == "Sum")
            & (pl.col("sk") == "Sluttbruker")
            & (pl.col("delar") == "Helår")
        )
        .with_columns(tilbyder=_provider)
        .filter(pl.col("tilbyder").is_not_null())
        .group_by("ar", "ms", "tilbyder")
        .agg(pl.col("svar").sum().alias("omsetning"))
    )
    arpu_provider = (
        _omsetning.join(_abonnement, on=["ar", "ms", "tilbyder"])
        .with_columns((pl.col("omsetning").cast(pl.Float64) * 1000 / pl.col("abonnement") / 12).alias("arpu"))
        .select("ar", "ms", "tilbyder", "arpu")
        .sort("ms", "tilbyder", "ar")
        .collect()
    )
    return (arpu_provider,)


@app.cell
def _(arpu_provider, mo, pl, plt, table_excel_download):
    _colors = {
        "Telenor": "#00B0F0",
        "Telia": "#9900CC",
        "Ice": "#ffc000",
        "Fjordkraft": "#385624",
        "Chili mobil": "#FF0000",
        "Plussmobil": "#A5A5A5",
        "Happybytes": "#698ED0",
        "Unifon": "#7F7F7F",
    }
    _private_order = ["Telenor", "Telia", "Ice", "Fjordkraft", "Chili mobil", "Plussmobil", "Happybytes"]
    _business_order = ["Telenor", "Telia", "Unifon", "Ice"]

    def _plot_arpu_provider(_segment, _order):
        _data = arpu_provider.filter(pl.col("ms") == _segment)
        _fig, _ax = plt.subplots(figsize=(7.2, 4.8), dpi=120)
        _label_offsets = {
            "Privat": {"Chili mobil": 18, "Plussmobil": 6, "Ice": -6, "Fjordkraft": -18},
            "Bedrift": {"Unifon": 10, "Telenor": -10, "Ice": 7, "Telia": -7},
        }.get(_segment, {})
        for _provider in _order:
            _provider_data = _data.filter(pl.col("tilbyder") == _provider).sort("ar")
            if _provider_data.is_empty():
                continue
            _ax.plot(
                _provider_data["ar"].to_list(),
                _provider_data["arpu"].to_list(),
                color=_colors[_provider],
                linewidth=2.6,
                solid_capstyle="round",
                label=_provider,
            )
            _ax.annotate(
                f"{_provider_data['arpu'][-1]:.0f}",
                xy=(_provider_data["ar"][-1], _provider_data["arpu"][-1]),
                xytext=(8, _label_offsets.get(_provider, 0)),
                textcoords="offset points",
                fontsize=8,
                color="#404040",
            )
        _ax.set_ylabel("NOK per måned")
        _ax.set_xticks(sorted(_data["ar"].unique().to_list()))
        _ax.grid(axis="y", color="#d9d9d9", linewidth=0.8)
        _ax.spines[["top", "right", "left"]].set_visible(False)
        _ax.spines["bottom"].set_color("#d9d9d9")
        _ax.tick_params(axis="both", colors="#555555", length=0)
        _ax.legend(loc="lower center", bbox_to_anchor=(0.5, -0.30), ncol=3, frameon=False, fontsize=8)
        _fig.tight_layout()
        return _fig

    _private_export = table_excel_download(
        arpu_provider.filter(pl.col("ms") == "Privat"),
        "figur-9-arpu-privat.xlsx",
        "Figur 9 - ARPU privatmarkedet",
        "ARPU privat",
    )
    _business_export = table_excel_download(
        arpu_provider.filter(pl.col("ms") == "Bedrift"),
        "figur-9-arpu-bedrift.xlsx",
        "Figur 9 - ARPU bedriftsmarkedet",
        "ARPU bedrift",
    )
    mo.hstack(
        [
            mo.vstack(
                [
                    mo.hstack([mo.Html('<div class="figure-heading-title">ARPU privatmarkedet</div>'), _private_export], justify="start", gap=0.35),
                    _plot_arpu_provider("Privat", _private_order),
                ],
                gap=0.5,
            ),
            mo.vstack(
                [
                    mo.hstack([mo.Html('<div class="figure-heading-title">ARPU bedriftsmarkedet</div>'), _business_export], justify="start", gap=0.35),
                    _plot_arpu_provider("Bedrift", _business_order),
                ],
                gap=0.5,
            ),
        ],
        justify="center",
        gap=2,
    )
    return


@app.cell
def _(mo):
    mo.Html(
        """
        <div style="
            width: 100%;
            margin: 46px 0 24px;
            border-top: 3px solid #0b2b66;
        "></div>
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ## 10 - Prisutvikling: omsetning per GB
    """)
    return


@app.cell
def _(df, pl):
    _provider_name = pl.col("fusnavn").str.to_lowercase()
    _total_provider = (
        pl.when(_provider_name.str.contains("telenor"))
        .then(pl.lit("Telenor"))
        .when(_provider_name.str.contains("telia"))
        .then(pl.lit("Telia"))
        .when(_provider_name.str.contains("lyse") | _provider_name.str.contains("ice"))
        .then(pl.lit("Ice"))
        .otherwise(pl.lit("Øvrige"))
    )
    _challenger_provider = (
        pl.when(_provider_name.str.contains("telenor"))
        .then(pl.lit("Telenor"))
        .when(_provider_name.str.contains("telia"))
        .then(pl.lit("Telia"))
        .when(_provider_name.str.contains("lyse") | _provider_name.str.contains("ice"))
        .then(pl.lit("Ice"))
        .when(_provider_name == "chili mobil")
        .then(pl.lit("Chili mobil"))
        .when(_provider_name == "happybytes")
        .then(pl.lit("Happybytes"))
        .when(_provider_name == "fjordkraft mobil")
        .then(pl.lit("Fjordkraft"))
        .when(_provider_name == "plussmobil")
        .then(pl.lit("Plussmobil"))
        .when(_provider_name == "unifon")
        .then(pl.lit("Unifon"))
        .otherwise(None)
    )

    def _nok_per_gb(_provider_expr):
        _inntekter = (
            df.filter(
                (pl.col("dk") == "Mobiltelefoni")
                & (pl.col("hg") == "Inntekter")
                & (pl.col("tp") == "Sum")
                & (pl.col("sk") == "Sluttbruker")
                & (pl.col("delar") == "Helår")
            )
            .with_columns(tilbyder=_provider_expr)
            .filter(pl.col("tilbyder").is_not_null())
            .group_by("ar", "tilbyder")
            .agg(pl.col("svar").sum().alias("inntekter"))
        )
        _trafikk = (
            df.filter(
                (pl.col("dk") == "Mobiltelefoni")
                & (pl.col("hg") == "Trafikk")
                & (pl.col("n1") == "Data")
                & (pl.col("tp") == "Sum")
                & (pl.col("sk") == "Sluttbruker")
                & (pl.col("delar") == "Helår")
            )
            .with_columns(tilbyder=_provider_expr)
            .filter(pl.col("tilbyder").is_not_null())
            .group_by("ar", "tilbyder")
            .agg(pl.col("svar").sum().alias("datatrafikk_gb"))
        )
        return (
            _inntekter.join(_trafikk, on=["ar", "tilbyder"])
            .with_columns(
                (
                    pl.col("inntekter").cast(pl.Float64)
                    * 1000
                    / pl.col("datatrafikk_gb")
                ).alias("nok_per_gb")
            )
            .select("ar", "tilbyder", "nok_per_gb")
            .sort("tilbyder", "ar")
            .collect()
        )

    nok_per_gb_total = _nok_per_gb(_total_provider)
    nok_per_gb_challengers = _nok_per_gb(_challenger_provider)
    return nok_per_gb_challengers, nok_per_gb_total


@app.cell
def _(
    mo,
    nok_per_gb_challengers,
    nok_per_gb_total,
    pl,
    plt,
    table_excel_download,
):
    _colors = {
        "Telenor": "#00B0F0",
        "Telia": "#9900CC",
        "Ice": "#ffc000",
        "Øvrige": "#00b050",
        "Chili mobil": "#FF0000",
        "Happybytes": "#698ED0",
        "Fjordkraft": "#385624",
        "Plussmobil": "#A5A5A5",
        "Unifon": "#7F7F7F",
    }

    def _spread_label_positions(_points, _minimum_gap):
        _ordered_points = sorted(_points, key=lambda _point: _point["y"])
        _label_y = {}
        _previous_y = None
        for _point in _ordered_points:
            _candidate_y = _point["y"]
            if _previous_y is not None:
                _candidate_y = max(_candidate_y, _previous_y + _minimum_gap)
            _label_y[_point["provider"]] = _candidate_y
            _previous_y = _candidate_y
        return _label_y

    def _plot_nok_per_gb(_data, _order, _legend_cols):
        _fig, _ax = plt.subplots(figsize=(7.2, 4.8), dpi=120)
        _endpoints = []
        for _provider in _order:
            _provider_data = _data.filter(pl.col("tilbyder") == _provider).sort("ar")
            if _provider_data.is_empty():
                continue
            _xs = _provider_data["ar"].to_list()
            _ys = _provider_data["nok_per_gb"].to_list()
            _ax.plot(
                _xs,
                _ys,
                color=_colors[_provider],
                linewidth=2.5,
                solid_capstyle="round",
                label=_provider,
            )
            _endpoints.append(
                {
                    "provider": _provider,
                    "x": _xs[-1],
                    "y": _ys[-1],
                    "label": f"{_ys[-1]:.1f}".replace(".", ","),
                }
            )
        _label_y = _spread_label_positions(
            _endpoints,
            2.4 if "Øvrige" not in _order else 1.8,
        )
        for _point in _endpoints:
            _ax.annotate(
                _point["label"],
                xy=(_point["x"], _point["y"]),
                xytext=(_point["x"] + 0.12, _label_y[_point["provider"]]),
                textcoords="data",
                fontsize=8,
                color="#404040",
                va="center",
            )
        _ax.set_ylabel("NOK per GB")
        _ax.set_xticks(sorted(_data["ar"].unique().to_list()))
        _ax.set_xlim(_data["ar"].min(), _data["ar"].max() + 0.35)
        _ax.grid(axis="y", color="#d9d9d9", linewidth=0.8)
        _ax.spines[["top", "right", "left"]].set_visible(False)
        _ax.spines["bottom"].set_color("#d9d9d9")
        _ax.tick_params(axis="both", colors="#555555", length=0)
        _ax.legend(loc="lower center", bbox_to_anchor=(0.5, -0.30), ncol=_legend_cols, frameon=False, fontsize=8)
        _fig.tight_layout()
        return _fig

    _total_export = table_excel_download(
        nok_per_gb_total,
        "figur-10-omsetning-per-gb-total.xlsx",
        "Figur 10 - Omsetning per GB totalt",
        "NOK per GB total",
    )
    _challenger_export = table_excel_download(
        nok_per_gb_challengers,
        "figur-10-omsetning-per-gb-tilbydere.xlsx",
        "Figur 10 - Omsetning per GB tilbydere",
        "NOK per GB tilbydere",
    )
    mo.hstack(
        [
            mo.vstack(
                [
                    mo.hstack([mo.Html('<div class="figure-heading-title">Totalt</div>'), _total_export], justify="start", gap=0.35),
                    _plot_nok_per_gb(nok_per_gb_total, ["Telenor", "Telia", "Ice", "Øvrige"], 4),
                ],
                gap=0.5,
            ),
            mo.vstack(
                [
                    mo.hstack([mo.Html('<div class="figure-heading-title">Utvalgte tilbydere</div>'), _challenger_export], justify="start", gap=0.35),
                    _plot_nok_per_gb(
                        nok_per_gb_challengers,
                        ["Telenor", "Telia", "Ice", "Chili mobil", "Happybytes", "Fjordkraft", "Plussmobil", "Unifon"],
                        4,
                    ),
                ],
                gap=0.5,
            ),
        ],
        justify="center",
        gap=2,
    )
    return


@app.cell
def _(mo):
    mo.Html(
        """
        <div style="
            width: 100%;
            margin: 46px 0 24px;
            border-top: 3px solid #0b2b66;
        "></div>
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ## 11 - Abonnementsandeler fra parquet
    """)
    return


@app.cell
def _(df, mo, pl, table_excel_download):
    _provider_name = pl.col("fusnavn").str.to_lowercase()
    _provider_group = (
        pl.when(_provider_name.str.contains("telenor"))
        .then(pl.lit("Telenor"))
        .when(_provider_name.str.contains("telia"))
        .then(pl.lit("Telia"))
        .when(
            _provider_name.is_in(["ice communication norge", "lyse tele"])
            | _provider_name.str.contains("lyse")
            | _provider_name.str.contains("ice")
        )
        .then(pl.lit("Lyse Tele (Ice)"))
        .otherwise(pl.lit("Øvrige"))
    )
    _shares = (
        df.filter(
            (pl.col("dk") == "Mobiltelefoni")
            & (pl.col("hg") == "Abonnement")
            & (pl.col("n1").is_in(["Fakturert", "Kontantkort"]))
            & (pl.col("n2") == "Ingen")
            & (pl.col("tp") == "Sum")
            & (pl.col("sk") == "Sluttbruker")
            & (pl.col("delar") == "Helår")
        )
        .with_columns(Grossist=_provider_group)
        .group_by("ar", "Grossist")
        .agg(pl.col("svar").sum().alias("abonnement"))
        .with_columns(
            markedsandel=pl.col("abonnement")
            / pl.col("abonnement").sum().over("ar")
        )
        .filter(pl.col("Grossist") != "Øvrige")
        .select(pl.col("ar").cast(pl.Utf8), "Grossist", "markedsandel")
        .collect()
    )
    _year_columns = sorted(_shares["ar"].unique().to_list())
    wholesale_market_shares = (
        _shares.pivot(
            values="markedsandel",
            index="Grossist",
            on="ar",
            aggregate_function="first",
        )
        .select(["Grossist"] + _year_columns)
        .sort("Grossist")
    )
    _year_columns = [column for column in wholesale_market_shares.columns if column != "Grossist"]
    _display = wholesale_market_shares.with_columns(
        [(pl.col(column) * 100).round(1).alias(column) for column in _year_columns]
    )
    _export = table_excel_download(
        wholesale_market_shares,
        "figur-11-wholesale-market-shares.xlsx",
        "Figur 11 - Abonnementsandeler fra parquet",
        "Wholesale",
        _percent_columns=_year_columns,
    )
    mo.vstack(
        [
            mo.hstack(
                [
                    mo.Html('<div class="figure-heading-title">Abonnementsandeler fra parquet</div>'),
                    _export,
                ],
                justify="start",
                gap=0.35,
            ),
            mo.ui.table(_display, page_size=3),
        ],
        gap=0.5,
    )
    return


@app.cell
def _(mo):
    mo.Html(
        """
        <div style="
            width: 100%;
            margin: 46px 0 24px;
            border-top: 3px solid #0b2b66;
        "></div>
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ## 12 - Markedskonsentrasjon
    """)
    return


if __name__ == "__main__":
    app.run()
